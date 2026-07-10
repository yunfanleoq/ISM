"""
ISM Automation — Excel 解析器
从 ISM 配置 Excel 中提取数据模型定义、设备清单、点位映射。

解析逻辑继承自 generate_ism_data.py，提取为可复用的模块。
"""
from __future__ import annotations

import re
from collections import OrderedDict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

from ism_automation.core.logger import ism_logger
from ism_automation.config.loader import ProjectConfig


class ExcelParser:
    """ISM Excel 配置解析器"""

    def __init__(self, config: ProjectConfig):
        self.config = config
        self.excel_path = Path(config.excel_path)
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel 文件不存在: {self.excel_path}")

        self.wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        ism_logger.info(f"📖 加载 Excel: {self.excel_path}")

    # ── 模板 Sheet 解析 ───────────────────────────────

    def parse_template(self) -> Dict[str, Dict[str, Any]]:
        """
        解析"模板"Sheet，提取数据模型定义（A20/A40/UPS）

        返回格式:
        {
            "A20电力仪表": {
                "model_name": "A20电力仪表 AI20 DI8",
                "ai_count": 20, "di_count": 8,
                "ai_points": [{"offset": 0, "name": "A相电压", "coeff": 0.1, "parse": 177}, ...],
                "di_points": [{"offset": 0, "name": "开关状态", "bit_offset": 0}, ...],
                "devices": [{"device_name": "1A1_U11_S18_1", "ai_start": 2000, "di_start": 3000}, ...],
            }
        }
        """
        ws = self.wb['模板']
        models = OrderedDict()
        current_model = None
        mapping = self.config.excel_mapping

        # 列映射（从配置读取，可覆盖）
        model_name_col = 0  # A列
        ai_name_col = 1     # B列
        ai_offset_col = 2   # C列
        ai_coeff_col = 3    # D列
        ai_parse_col = 4    # E列
        di_offset_col = 8   # I列
        di_name_col = 9     # J列
        di_bit_col = 10     # K列
        dev_name_col = 11   # L列
        ai_start_col = 14   # O列
        di_start_col = 15   # P列

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            col0 = str(row[0]).strip() if row[0] else ''

            # 检测模型标题行（A20/A40/UPS/施耐德）
            if col0 and ('A20' in col0 or 'A40' in col0 or 'UPS' in col0 or '施耐德' in col0):
                current_model = col0.replace('\n', ' ')
                ai_count = 0
                di_count = 0
                m_ai = re.search(r'AI\s*(\d+)', current_model)
                m_di = re.search(r'DI\s*(\d+)', current_model)
                if m_ai:
                    ai_count = int(m_ai.group(1))
                if m_di:
                    di_count = int(m_di.group(1))

                # 标准化模型 key
                if 'A20' in current_model:
                    model_key = 'A20电力仪表'
                elif 'A40' in current_model:
                    model_key = 'A40电力仪表'
                else:
                    model_key = '施耐德UPS'

                models[model_key] = {
                    'model_name': current_model,
                    'ai_count': ai_count,
                    'di_count': di_count,
                    'ai_points': [],
                    'di_points': [],
                    'devices': [],
                }
                continue

            if not current_model:
                continue

            # 标准化模型 key
            if 'A20' in current_model:
                model_key = 'A20电力仪表'
            elif 'A40' in current_model:
                model_key = 'A40电力仪表'
            else:
                model_key = '施耐德UPS'

            if model_key not in models:
                continue

            # 解析 AI 点
            ai_name = str(row[ai_name_col]).strip() if row[ai_name_col] else ''
            ai_offset = row[ai_offset_col]
            ai_coeff = row[ai_coeff_col]
            ai_parse = row[ai_parse_col]

            # 解析 DI 点
            di_name = str(row[di_name_col]).strip() if row[di_name_col] else ''
            di_offset = row[di_offset_col]
            di_bit = row[di_bit_col]

            # 解析设备信息
            dev_name = str(row[dev_name_col]).strip() if row[dev_name_col] else ''
            ai_start = row[ai_start_col] if row[ai_start_col] else None
            di_start = row[di_start_col] if row[di_start_col] else None

            # 添加 AI 点
            if ai_name and ai_name not in ['None', '']:
                coeff_val = ai_coeff
                parse_val = ai_parse
                # 检测 coeff/parse 互换（A40 特殊处理）
                if isinstance(ai_coeff, (int, float)) and isinstance(ai_parse, (int, float)):
                    if ai_coeff > 50 and ai_parse < 1:
                        parse_val = ai_coeff
                        coeff_val = ai_parse

                models[model_key]['ai_points'].append({
                    'offset': int(ai_offset) if ai_offset is not None else None,
                    'name': ai_name,
                    'coeff': coeff_val if coeff_val else 1,
                    'parse': int(parse_val) if parse_val else 177,
                })

            # 添加 DI 点
            if di_name and di_name not in ['None', '']:
                models[model_key]['di_points'].append({
                    'offset': int(di_offset) if di_offset is not None else None,
                    'name': di_name,
                    'bit_offset': int(di_bit) if di_bit is not None else None,
                })

            # 添加设备
            if dev_name and dev_name not in ['None', '']:
                models[model_key]['devices'].append({
                    'device_name': dev_name,
                    'ai_start': int(ai_start) if ai_start else None,
                    'di_start': int(di_start) if di_start else None,
                })

        # 过滤掉 ai_start 为空的设备
        for key in list(models.keys()):
            models[key]['devices'] = [
                d for d in models[key]['devices']
                if d['ai_start'] is not None
            ]

        ism_logger.info(f"📊 解析模板: {len(models)} 个模型, "
                       f"{sum(len(m['ai_points']) for m in models.values())} 个 AI 点, "
                       f"{sum(len(m['di_points']) for m in models.values())} 个 DI 点, "
                       f"{sum(len(m['devices']) for m in models.values())} 个设备")
        return models

    # ── 主数据 Sheet 解析 ─────────────────────────────

    def parse_main_data(self) -> List[Dict[str, Any]]:
        """
        解析主数据 Sheet，提取所有寄存器映射记录

        返回格式:
        [
            {"point_no": 1, "node_id": 1, "node_point": 1, "node_name": "", "point_name": "A相电压", "reg_addr": 0},
            ...
        ]
        """
        # 主数据 Sheet 名称通常是文件名本身
        sheet_name = self.excel_path.stem
        if sheet_name not in self.wb.sheetnames:
            # 尝试第一个非特殊 Sheet
            sheet_name = self.wb.sheetnames[0] if self.wb.sheetnames else None

        if not sheet_name:
            ism_logger.warning("未找到主数据 Sheet")
            return []

        ws = self.wb[sheet_name]
        records = []

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            point_no = row[0]
            node_id = row[1]
            node_point = row[2]
            node_name = str(row[3]).strip() if row[3] else ''
            point_name = str(row[4]).strip() if row[4] else ''
            reg_addr = row[5]

            # 跳过非数据行
            if point_no is None or not str(point_no).strip():
                continue
            try:
                pn = int(float(str(point_no)))
            except (ValueError, TypeError):
                continue

            try:
                ni = int(float(str(node_id))) if node_id is not None and str(node_id).strip() else None
            except (ValueError, TypeError):
                ni = str(node_id) if node_id else None

            try:
                np_val = int(float(str(node_point))) if node_point is not None and str(node_point).strip() else None
            except (ValueError, TypeError):
                np_val = None

            try:
                ra = int(float(str(reg_addr))) if reg_addr is not None and str(reg_addr).strip() else None
            except (ValueError, TypeError):
                ra = None

            records.append({
                'point_no': pn,
                'node_id': ni,
                'node_point': np_val,
                'node_name': node_name,
                'point_name': point_name,
                'reg_addr': ra,
            })

        ism_logger.info(f"📊 解析主数据: {len(records)} 条记录")
        return records

    # ── 设备清单解析 ──────────────────────────────────

    def parse_device_list(self) -> List[Dict[str, Any]]:
        """
        解析 Sheet1 和 Sheet3，获取完整设备清单

        返回格式:
        [
            {"full_name": "1A1_U11_S18_1", "short_name": "1A1_S18_1", "ai_start": 2000, "di_start": 3000},
            ...
        ]
        """
        devices_full = []

        # Sheet1
        if 'Sheet1' in self.wb.sheetnames:
            ws1 = self.wb['Sheet1']
            for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row, values_only=True):
                full_name = str(row[0]).strip() if row[0] else ''
                short_name = str(row[1]).strip() if row[1] else ''
                ai_start = row[2] if row[2] else None
                if full_name and short_name:
                    devices_full.append({
                        'full_name': full_name,
                        'short_name': short_name,
                        'ai_start': int(ai_start) if ai_start else None,
                        'di_start': None,
                    })

        # Sheet3（DI 起始地址）
        if 'Sheet3' in self.wb.sheetnames:
            ws3 = self.wb['Sheet3']
            sheet3_data = {}
            for row in ws3.iter_rows(min_row=1, max_row=ws3.max_row, values_only=True):
                dev_name = str(row[0]).strip() if row[0] else ''
                ai_start = row[1] if row[1] else None
                di_start = row[2] if row[2] else None
                if dev_name:
                    sheet3_data[dev_name] = {
                        'ai_start': int(ai_start) if ai_start else None,
                        'di_start': int(di_start) if di_start else None,
                    }

            # 匹配 DI 起始地址
            for dev in devices_full:
                for s3_name, s3_info in sheet3_data.items():
                    if dev['full_name'] in s3_name or s3_name.startswith(dev['full_name']):
                        dev['di_start'] = s3_info['di_start']
                        break

        ism_logger.info(f"📊 解析设备清单: {len(devices_full)} 台设备")
        return devices_full

    # ── 辅助工具 ──────────────────────────────────────

    @staticmethod
    def determine_device_type(full_name: str) -> str:
        """根据设备名称判断设备类型"""
        if 'A40' in full_name.lower() or '_A40' in full_name:
            return 'A40电力仪表'
        if 'UPS' in full_name.lower() or '施耐德' in full_name:
            return '施耐德UPS'
        return 'A20电力仪表'

    @staticmethod
    def get_unit_for_point(point_name: str) -> str:
        """根据点位名称推断单位"""
        units = {
            '电压': 'V', '电流': 'A', '功率': 'kW',
            '频率': 'Hz', '功率因数': '', '电能': 'kWh',
            '温度': '°C', '湿度': '%',
        }
        for key, unit in units.items():
            if key in point_name:
                return unit
        return ''

    # ── 完整解析 ─────────────────────────────────────

    def parse_all(self) -> Dict[str, Any]:
        """解析所有 Sheet，返回完整结构化数据"""
        return {
            'template': self.parse_template(),
            'main_data': self.parse_main_data(),
            'device_list': self.parse_device_list(),
        }
