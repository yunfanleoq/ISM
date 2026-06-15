#!/usr/bin/env python3
"""
ISM Excel 点位表导入脚本（通用版）
用法: python3 generate_ism_data.py <Excel文件路径> [--project-name 项目名]
"""
import openpyxl, json, re, os, sys, argparse, uuid
from collections import OrderedDict

def get_parse_label(c): return {177:'uint16',179:'int32',73:'int16',71:'uint16'}.get(c,f'code_{c}')
def get_unit_for_point(n):
    if '电压' in n: return 'V'
    if '电流' in n and '谐波' not in n: return 'A'
    if '功率' in n:
        if '因数' in n: return ''; return {'有功':'kW','无功':'kvar','视在':'kVA'}.get(next((k for k in ['有功','无功','视在'] if k in n),''),'kW')
    if '频率' in n: return 'Hz'
    if '电度' in n: return 'kWh'
    if '谐波畸变率' in n: return '%'
    if '温度' in n: return '°C'
    if '时间' in n: return 'min'
    return ''
def determine_device_type(n):
    if 'A40' in n.lower() or '_A40' in n: return 'A40电力仪表'
    if 'UPS' in n.lower() or '施耐德' in n: return '施耐德UPS'
    return 'A20电力仪表'
def infer_location(s):
    loc={'building':'配电室','cabinet':'','unit':''}
    for p,h in [(r'(\w+\d)_U(\d+)_S(\d+)_(\d+)','U{n1}_S{n2}'),(r'(\w+\d)_U(\d+)_D(\d+)_(\d+)','U{n1}_D{n2}'),(r'UPS_(\w+\d)_U(\d+)','UPS_U{n1}')]:
        m=re.search(p,s)
        if m:
            loc['building']=f"{m.group(1)}区"
            loc['cabinet']=h.format(n1=m.group(2),n2=m.group(3) if len(m.groups())>3 else m.group(2))
            loc['unit']=m.group(m.lastindex or 1);break
    loc['description']=f"{loc['building']}, {loc['cabinet']}";return loc

def detect_sheets(wb):
    r={k:None for k in ['main','template','device1','device3']}
    for sn in wb.sheetnames:
        ws=wb[sn]
        if ws.max_row<3 or ws.max_column<3:continue
        # Check first 3 rows for detection keywords (headers may be in row 3)
        hlines=[]
        for rn in range(1,min(4,ws.max_row+1)):
            hlines.append(' '.join([str(c.value)[:60] if c.value else '' for c in ws[rn]]))
        hall=' '.join(hlines)
        if '模版类型' in hall or 'AI名称' in hall: r['template']=sn;continue
        if '点号' in hall or ('寄存器' in hall and '地址' in hall) or '源节点' in hall: r['main']=sn;continue
        # Check if this sheet looks like a main data table (many rows, 7+ columns, numeric data)
        if ws.max_row>100 and ws.max_column>=5:
            # Check row 4-6 for numeric point numbers
            has_numeric=False
            for rn in range(4,min(8,ws.max_row+1)):
                try:
                    v=ws.cell(rn,1).value
                    if isinstance(v,(int,float)) or (isinstance(v,str) and re.match(r'^\d+$',v.strip())):
                        has_numeric=True;break
                except:pass
            if has_numeric: r['main']=sn;continue
        if ws.max_column<=3 and ws.max_row>10:
            v2=ws.cell(2,1).value
            if isinstance(v2,str) and re.search(r'P\d+|U\d+_[SD]\d+|UPS',v2):
                if isinstance(ws.cell(2,3).value,(int,float)) and int(ws.cell(2,3).value)>100: r['device1']=sn;continue
                if isinstance(ws.cell(2,2).value,(int,float)) and int(ws.cell(2,2).value)>100: r['device3']=sn;continue
    # Fallback: if main not found, use the largest sheet (most rows)
    if not r['main']:
        largest=(None,0)
        for sn in wb.sheetnames:
            ws=wb[sn]
            if ws.max_row>largest[1] and ws.max_column>=5: largest=(sn,ws.max_row)
        r['main']=largest[0]
    return r

def parse_template(wb,sn):
    if not sn or sn not in wb.sheetnames: return OrderedDict()
    ws=wb[sn]; models=OrderedDict(); cur=None
    for row in ws.iter_rows(min_row=1,max_row=ws.max_row,values_only=True):
        c0=str(row[0]).strip() if row[0] else ''
        if c0 and any(x in c0 for x in ['A20','A40','UPS','施耐德']):
            cur = 'A20电力仪表' if 'A20' in c0 else ('A40电力仪表' if 'A40' in c0 else '施耐德UPS')
            ai_n=di_n=0
            for m in re.finditer(r'(AI|DI)\s*(\d+)',c0): 
                if m.group(1)=='AI':ai_n=int(m.group(2))
                else:di_n=int(m.group(2))
            models[cur]={'model_name':c0.replace('\n',' '),'ai_count':ai_n,'di_count':di_n,'ai_points':[],'di_points':[],'devices':[]}
            continue
        if not cur or cur not in models: continue
        ai_name=str(row[1]).strip() if row[1] else ''
        ai_off=row[2];ai_coeff=row[3];ai_parse=row[4]
        di_name=str(row[8]).strip() if row[8] else ''
        di_off=row[9];di_bit=row[10]
        dv_name=str(row[11]).strip() if row[11] else ''
        ai_start=row[14];di_start=row[15]
        if ai_name and ai_name!='None':
            cv, pv = ai_coeff, ai_parse
            if isinstance(ai_coeff,(int,float)) and isinstance(ai_parse,(int,float)) and ai_coeff>50 and ai_parse<1: pv, cv = ai_coeff, ai_parse
            models[cur]['ai_points'].append({'offset':int(ai_off) if ai_off is not None else None,'name':ai_name,'coeff':cv if cv else 1,'parse':int(pv) if pv else 177})
        if di_name and di_name!='None':
            models[cur]['di_points'].append({'offset':int(di_off) if di_off is not None else None,'name':di_name,'bit_offset':int(di_bit) if di_bit is not None else None})
        if dv_name and dv_name!='None':
            models[cur]['devices'].append({'device_name':dv_name,'ai_start':int(ai_start) if ai_start else None,'di_start':int(di_start) if di_start else None})
    for k in list(models.keys()): models[k]['devices']=[d for d in models[k]['devices'] if d['ai_start'] is not None]
    return models

def parse_devices(wb,s1,s3):
    devs=[]
    if s1 and s1 in wb.sheetnames:
        for row in wb[s1].iter_rows(min_row=1,max_row=wb[s1].max_row,values_only=True):
            fn=str(row[0]).strip() if row[0] else ''; sn=str(row[1]).strip() if row[1] else ''; a=row[2]
            if fn and sn: devs.append({'full_name':fn,'short_name':sn,'ai_start':int(a) if a else None,'di_start':None})
    if s3 and s3 in wb.sheetnames:
        s3d={}
        for row in wb[s3].iter_rows(min_row=1,max_row=wb[s3].max_row,values_only=True):
            dn=str(row[0]).strip() if row[0] else ''; a=row[1]; d=row[2]
            if dn: s3d[dn]={'ai_start':int(a) if a else None,'di_start':int(d) if d else None}
        for dv in devs:
            for k,v in s3d.items():
                if dv['full_name'] in k or k.startswith(dv['full_name']): dv['di_start']=v['di_start'];break
    return devs

def parse_main(wb,sn):
    ws=wb[sn]; recs=[]
    for row in ws.iter_rows(min_row=1,max_row=ws.max_row,values_only=True):
        pn=row[0]; ni=row[1]; np_=row[2]; nn=str(row[3]).strip() if row[3] else ''; pt=str(row[4]).strip() if row[4] else ''; ra=row[5]
        if pn is None or not str(pn).strip(): continue
        try: pn_v=int(float(str(pn)))
        except: continue
        try: ni_v=int(float(str(ni))) if ni is not None and str(ni).strip() else None
        except: ni_v=str(ni) if ni else None
        try: np_v=int(float(str(np_))) if np_ is not None and str(np_).strip() else None
        except: np_v=None
        try: ra_v=int(float(str(ra))) if ra is not None and str(ra).strip() else None
        except: ra_v=None
        recs.append({'point_no':pn_v,'node_id':ni_v,'node_point':np_v,'node_name':nn,'point_name':pt,'reg_addr':ra_v})
    return recs

def extract_gateway(wb,sn):
    ip='192.168.1.1';port=502;ip2=''
    try:
        ws=wb[sn]
        for r in [ws[1],ws[2]]:
            for c in r:
                s=str(c.value) if c.value else ''
                for m in re.finditer(r'(\d+\.\d+\.\d+\.\d+)',s):
                    if ip=='192.168.1.1':ip=m.group(1)
                    elif not ip2:ip2=m.group(1)
                m=re.search(r'端口.*?(\d+)',s) or re.search(r'port.*?(\d+)',s,re.I)
                if m and port==502: port=int(m.group(1))
    except: pass
    return ip,port,ip2

def main():
    p=argparse.ArgumentParser(description='ISM Excel点位表导入',epilog='示例: python3 generate_ism_data.py "配电室.xlsx"')
    p.add_argument('excel_path');p.add_argument('--project-name','-n');p.add_argument('--main-sheet','-m');p.add_argument('--template-sheet','-t');p.add_argument('--device1-sheet','-d1');p.add_argument('--device3-sheet','-d3');p.add_argument('--output-dir','-o',default='.');p.add_argument('--dry-run',action='store_true')
    args=p.parse_args()
    if not os.path.exists(args.excel_path): print(f"文件不存在: {args.excel_path}");sys.exit(1)
    pname=args.project_name or re.sub(r'\d+\.\d+\.\d+\.\d+','',os.path.splitext(os.path.basename(args.excel_path))[0]).strip() or '配电室项目'
    print(f"ISM Excel导入 | 项目: {pname} | 文件: {args.excel_path}")
    wb=openpyxl.load_workbook(args.excel_path,data_only=True)
    sc=detect_sheets(wb)
    ms=args.main_sheet or sc['main'];ts=args.template_sheet or sc['template'];d1=args.device1_sheet or sc['device1'];d3=args.device3_sheet or sc['device3']
    if args.dry_run:
        print(f"Sheet: 主={ms} 模板={ts} 设备1={d1} 设备3={d3}")
        return
    ip,port,ip2=extract_gateway(wb,ms)
    print(f"[1/4] 解析模板...")
    models=parse_template(wb,ts)
    for k,m in models.items(): print(f"  {k}: AI={len(m['ai_points'])} DI={len(m['di_points'])}")
    print(f"[2/4] 解析设备...")
    devs=parse_devices(wb,d1,d3);print(f"  {len(devs)}台")
    print(f"[3/4] 解析寄存器映射...")
    recs=parse_main(wb,ms);print(f"  {len(recs)}条")
    rbn=OrderedDict()
    for r in recs:
        if r['node_name'] not in rbn: rbn[r['node_name']]=[]
        rbn[r['node_name']].append(r)
    dbg=OrderedDict();tp=0
    for dv in devs:
        mt=determine_device_type(dv['full_name']);gm=re.match(r'(P\d+)',dv['full_name']);g=gm.group(1) if gm else '管理机'
        if g not in dbg: dbg[g]=[]
        dbg[g].append({**dv,'model_type':mt,'data_points':[]})
    for g,ds in dbg.items():
        cands=[]
        for nn,rs in rbn.items():
            if g in nn: cands=rs;break
        if not cands: continue
        ri=0
        for d in ds:
            m=models.get(d['model_type'])
            if not m: continue
            npt=len(m['ai_points'])+len(m['di_points']);dr=cands[ri:ri+npt]
            if not dr: continue
            ri+=npt;di=ds.index(d)
            for pi,rr in enumerate(dr):
                if pi<len(m['ai_points']): mp=m['ai_points'][pi];tt='AI'
                elif pi<len(m['ai_points'])+len(m['di_points']): mp=m['di_points'][pi-len(m['ai_points'])];tt='DI'
                else: mp=None;tt='AI'
                mn=mp['name'] if mp else rr['point_name']
                is_al=('通讯状态' in mn)
                dbg[g][di]['data_points'].append({'name':rr['point_name'],'display_name':mn,'model_name':mn,'point_type':tt,'register_addr':rr['reg_addr'],'model_offset':mp['offset'] if mp else None,'coeff':mp.get('coeff',1) if mp else 1,'parse':mp.get('parse',177) if mp else 177,'unit':get_unit_for_point(mn),'is_alarm':is_al,'alarm_level':3 if is_al else 0,'alarm_message':'设备通讯离线' if is_al else ''})
            tp+=len(dr)
    print(f"  {sum(len(d) for d in dbg.values())}台设备, {tp}个数据点")
    print(f"[4/4] 生成文件...")
    os.makedirs(args.output_dir,exist_ok=True)
    sn=re.sub(r'[^\w\u4e00-\u9fff\-]','_',pname);pf=os.path.join(args.output_dir,sn)
    # analysis.json
    an=OrderedDict();an["project_name"]=pname;an["protocol"]="Modbus TCP";an["gateway"]={"primary_ip":ip,"port":port}
    if ip2: an["gateway"]["secondary_ip"]=ip2
    an["total_devices"]=len(devs);an["total_data_points"]=tp;an["device_categories"]=OrderedDict()
    for ck,m in models.items():
        ct=OrderedDict();ct["description"]=f"{len(m['ai_points'])}AI+{len(m['di_points'])}DI";ct["ai_count"]=len(m['ai_points']);ct["di_count"]=len(m['di_points']);ct["register_offset_per_device"]=len(m['ai_points'])+len(m['di_points']);ct["devices"]=[]
        ct["data_point_types"]=OrderedDict()
        ct["data_point_types"]["AI"]=[OrderedDict([("offset",pt["offset"]),("name",pt["name"]),("coeff",pt["coeff"]),("parse",pt["parse"]),("parse_label",get_parse_label(pt["parse"])),("unit",get_unit_for_point(pt["name"]))]) for pt in m['ai_points']]
        ct["data_point_types"]["DI"]=[OrderedDict([("offset",pt["offset"]),("name",pt["name"]),("bit_offset",pt["bit_offset"])]) for pt in m['di_points']]
        for g,ds in dbg.items():
            for d in ds:
                if d["model_type"]==ck: ct["devices"].append(OrderedDict([("group",g),("full_name",d["full_name"]),("short_name",d["short_name"]),("ai_start",d["ai_start"]),("di_start",d["di_start"]),("data_point_count",len(d["data_points"]))]))
        an["device_categories"][ck]=ct
    with open(f"{pf}_analysis.json",'w',encoding='utf-8') as f: json.dump(an,f,ensure_ascii=False,indent=2)
    print(f"  ✓ {sn}_analysis.json")
    # device_points.json
    dp=OrderedDict();dp["project"]=pname;dp["gateway_ip"]=ip;dp["gateway_port"]=port;dp["total_devices"]=sum(len(d) for d in dbg.values());dp["total_data_points"]=tp;dp["device_groups"]=OrderedDict()
    for g,ds in dbg.items():
        allp=[p for d in ds for p in d["data_points"]]
        dp["device_groups"][g]=OrderedDict([("model_type",ds[0]["model_type"] if ds else "Unknown"),("device_count",len(ds)),("point_count",len(allp)),("devices",[OrderedDict([("full_name",d["full_name"]),("short_name",d["short_name"]),("ai_start",d["ai_start"]),("di_start",d["di_start"]),("data_points",d["data_points"])]) for d in ds])])
    with open(f"{pf}_device_points.json",'w',encoding='utf-8') as f: json.dump(dp,f,ensure_ascii=False,indent=2)
    print(f"  ✓ {sn}_device_points.json")
    # 先生成组态大屏 UUID（后续构造数据模型和设备实例时直接填入）
    _display_model_uuid = str(uuid.uuid4())
    _display_uid = str(uuid.uuid4())
    _page_id = str(uuid.uuid4())

    # ISM package (simplified) —— 构造时即带 configUid/PageUUID，无需事后回填
    dms=[]
    for ck,m in models.items():
        rgs=[]
        ai_a=[OrderedDict([("name",pt["name"]),("offset",pt["offset"]),("register_type","holding"),("data_type",get_parse_label(pt["parse"])),("unit",get_unit_for_point(pt["name"])),("coeff",pt["coeff"]),("conversion_expression",f"x*{pt['coeff']}"),("is_alarm",False),("is_record",True),("RecordType",1),("RecordInterval",5),("RecordDataCharge","0"),("RecordDataTimely","0"),("FloatAccuracy",str(float(pt["coeff"])))]) for pt in m['ai_points']]
        di_a=[OrderedDict([("name",pt["name"]),("offset",pt["offset"]),("register_type","coil"),("data_type","bit"),("unit",""),("coeff",1),("is_alarm","通讯状态" in pt["name"]),("alarm_level",3 if "通讯状态" in pt["name"] else 0),("alarm_message","设备通讯离线" if "通讯状态" in pt["name"] else ""),("RecordType",1),("RecordInterval",5),("RecordDataCharge","0"),("RecordDataTimely","0"),("FloatAccuracy","0")]) for pt in m['di_points']]
        rgs.append(OrderedDict([("group_name","AI模拟量"),("register_type","holding"),("addresses",ai_a)]))
        rgs.append(OrderedDict([("group_name","DI数字量"),("register_type","coil"),("addresses",di_a)]))
        dms.append(OrderedDict([("name",ck),("protocol","modbus"),("configUid",_display_uid),("PageUUID",_page_id),("connection",OrderedDict([("ip",ip),("port",port),("slave_id",1)])),("register_groups",rgs)]))
    ism_ds=[]
    for g,ds in dbg.items():
        for d in ds:
            loc=infer_location(d["short_name"])
            ism_ds.append(OrderedDict([("name",d["short_name"]),("display_name",d["full_name"]),("model_name",d["model_type"]),("group",g),("register_offset",d["ai_start"]),("di_start",d["di_start"]),("location",loc),("data_point_count",len(d["data_points"])),("configUid",_display_uid),("PageUUID",_page_id),("offlineClear",0),("offlineDefaultValue","0"),("timeout",5),("interval",5000),("packTime",5000),("failedTimes",5),("IsEnable",1)]))
    alms=[OrderedDict([("name",f"{d['name']}_通讯离线告警"),("device_name",d["name"]),("data_point","主通讯状态"),("condition","x == 0"),("level",3),("keep_time",10),("alarm_message",f"{d['name']} 设备通讯离线")]) for d in ism_ds]
    c1=OrderedDict();c3=OrderedDict()
    for g,ds in dbg.items():
        for d in ds:
            loc=infer_location(d["short_name"]);cab=loc["cabinet"]
            if "1A3" in loc["building"]:
                if cab not in c3: c3[cab]=[]
                c3[cab].append(d)
            else:
                if cab not in c1: c1[cab]=[]
                c1[cab].append(d)
    td=sum(len(ds) for ds in dbg.values())
    dash=OrderedDict([("name",f"{pname}监控总览"),("size","1920x1080"),("style",OrderedDict([("theme","科技蓝"),("primary_color","#00d4ff"),("bg_color","#0a1628")])),("overview_page",OrderedDict([("name","总览"),("is_home",True),("sections",[OrderedDict([("name","顶部"),("zone","top"),("height",80),("component",OrderedDict([("type","DvBorderBox1"),("text",f"{pname} - 监控总览 ({td}台设备)")]))]),OrderedDict([("name","左侧设备树"),("zone","left"),("width",280),("component",OrderedDict([("type","DeviceTree"),("title","设备拓扑")]))]),OrderedDict([("name","区域1配电柜"),("zone","center_top"),("height",480),("cabinets",[OrderedDict([("name",cn),("device_count",len(cd)),("devices",[OrderedDict([("name",dd["short_name"]),("full_name",dd["full_name"]),("model_type",dd["model_type"]),("key_data",["AB线电压","A相电流","总有功功率","主通讯状态"]),("interactions",{"hover":{"show":f"detail_popup_{dd['short_name']}"}})]) for dd in cd])]) for cn,cd in c1.items()])]),OrderedDict([("name","区域2配电柜"),("zone","center_bottom"),("height",480),("cabinets",[OrderedDict([("name",cn),("device_count",len(cd)),("devices",[OrderedDict([("name",dd["short_name"]),("full_name",dd["full_name"]),("model_type",dd["model_type"]),("key_data",["AB线电压","A相电流","总有功功率","主通讯状态"])]) for dd in cd])]) for cn,cd in c3.items()])]),OrderedDict([("name","右侧告警"),("zone","right"),("width",660),("components",[OrderedDict([("type","alarmList"),("title","实时告警")]),OrderedDict([("type","RealDataTable"),("title","关键数据")])])])])])),("interaction_templates",OrderedDict([("hover_tooltip",OrderedDict([("trigger","mouseenter"),("action","visible"),("showItems","[detail_popup_{{device_name}}]"),("hide",OrderedDict([("trigger","mouseleave"),("action","visible"),("hideItems","[detail_popup_{{device_name}}]")]))]))]))])
    pkg=OrderedDict([("_meta",OrderedDict([("format_version","1.0"),("generator","ISM Excel Import Skill v2"),("project_name",pname)])),("project",OrderedDict([("name",pname),("description",f"{pname} Modbus TCP监控"),("gateway",OrderedDict([("primary_ip",ip),("port",port)]))])),("data_models",dms),("devices",ism_ds),("alarm_triggers",alms),("displayModel",OrderedDict([("uuid",_display_model_uuid),("name",f"{pname}监控大屏"),("project_uuid",""),("description",f"{pname}电力监控主屏"),("displayUid",_display_uid),("DisplayImage",""),("DisplayUserList",""),("DisplayType",1)])),("displayLayer",OrderedDict([("modelId",_display_uid),("pageName","主页面"),("pageId",_page_id),("isHome",1),("isLogin",0),("pageType",1),("layer",OrderedDict([("backColor","#f0f2f5"),("backgroundImage",""),("widthHeightRatio","16:9"),("width",1920),("height",1080)])),("components",OrderedDict([("cells",[])])),("_note","空大屏骨架，导入后在 ISMDisPlay 编辑器中拖拽组件填充")])),("dashboard",dash),("import_guide",["1.创建项目","2.导入数据模型（已含组态大屏关联）","3.导入设备","4.导入告警触发器","5.生成/导入组态大屏"])])
    with open(f"{pf}_ISM项目包.json",'w',encoding='utf-8') as f: json.dump(pkg,f,ensure_ascii=False,indent=2)
    print(f"  ✓ {sn}_ISM项目包.json")
    print("\n"+"="*60);print(f"完成！输出目录: {args.output_dir}");print("="*60)

if __name__=="__main__": main()
