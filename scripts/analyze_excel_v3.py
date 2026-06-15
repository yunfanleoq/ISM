#!/usr/bin/env python3
"""
ISM Excel 数据点分析器 - 最终修正版
正确理解模板表结构（模板类型+设备实例混合表），提取完整数据模型
"""
import openpyxl
import json
import re
from collections import defaultdict
from pathlib import Path


def analyze_excel(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # ========== 1. 读取主表（数据点列表）==========
    main_sheet = wb['1A配电室 172.31.4.14 172.20.255.14']
    all_rows = list(main_sheet.iter_rows(min_row=1, values_only=True))
    
    # 找到表头行
    header_row_idx = None
    for i, row in enumerate(all_rows):
        if row[0] and str(row[0]).strip() == '点号':
            header_row_idx = i
            break
    
    headers = [str(c).strip() if c else '' for c in all_rows[header_row_idx]]
    print(f"[主表] 表头: {headers}")
    
    # 读取数据行
    data_rows = []
    for i in range(header_row_idx + 1, len(all_rows)):
        row = all_rows[i]
        if row[0] is not None and str(row[0]).strip() != '' and str(row[0]).strip() != '点号':
            data_rows.append(row)
    
    print(f"[主表] 数据行数: {len(data_rows)}")
    
    # 按源节点分组
    device_points = defaultdict(list)
    for row in data_rows:
        src_name = str(row[3]).strip() if len(row) > 3 and row[3] else ''
        yx_name = str(row[4]).strip() if len(row) > 4 and row[4] else ''
        reg_addr = row[5] if len(row) > 5 and row[5] is not None else ''
        
        if src_name and src_name != '源节点名':
            device_points[src_name].append({
                'yx_name': yx_name,
                'reg_addr': int(reg_addr) if isinstance(reg_addr, (int, float)) else str(reg_addr)
            })
    
    print(f"[主表] 源节点数: {len(device_points)}")
    for name, pts in sorted(device_points.items(), key=lambda x: -len(x[1])):
        print(f"  {name}: {len(pts)} 个数据点")
    
    # ========== 2. 读取模板表（模板类型+设备实例）==========
    template_sheet = wb['模板']
    tpl_rows = list(template_sheet.iter_rows(min_row=1, values_only=True))
    
    # 模板表结构：
    # 列A: 模板类型（A20 AI 30个寄存器 DI 3个寄存器）
    # 列B: AI名称（AB线电压、BC线电压...）
    # 列C: 偏移（0, 2, 4...）
    # 列D: 系数（0.01, 0.001...）
    # 列E: 解析方式（177, 179...）
    # 列I: DI名称（输入状态2...）
    # 列J: 寄存器偏移（1, 2...）
    # 列K: 位偏移
    # 列L: 设备名称（1A1_U11_S18_1...）
    # 列M: 组态名称
    # 列N: AI起始地址
    # 列O: DI起始地址
    # 列R: 转发顺序
    # 列S: AI（AI起始地址）
    # 列T: DI（DI起始地址）
    
    print(f"\n[模板] 表头: {[str(c) if c else '' for c in tpl_rows[0]]}")
    
    # 解析模板表
    templates = []  # 模板类型列表
    current_template = None
    devices_from_template = []  # 从模板表提取的设备实例
    
    for i, row in enumerate(tpl_rows[1:], 1):  # 跳过表头
        col_a = str(row[0]).strip() if row[0] else ''
        ai_name = str(row[1]).strip() if len(row) > 1 and row[1] else ''
        offset = row[2] if len(row) > 2 and row[2] is not None else ''
        factor = row[3] if len(row) > 3 and row[3] is not None else ''
        parse_mode = row[4] if len(row) > 4 and row[4] is not None else ''
        di_name = str(row[8]).strip() if len(row) > 8 and row[8] else ''
        di_reg_offset = row[9] if len(row) > 9 and row[9] is not None else ''
        di_bit_offset = row[10] if len(row) > 10 and row[10] is not None else ''
        device_name = str(row[11]).strip() if len(row) > 11 and row[11] else ''
        ai_start_addr = row[13] if len(row) > 13 and row[13] is not None else ''
        di_start_addr = row[14] if len(row) > 14 and row[14] is not None else ''
        
        # 检测模板类型行（包含 "AI" 和 "DI" 说明）
        if 'AI' in col_a and 'DI' in col_a and ('寄存器' in col_a or '个' in col_a):
            # 提取模板类型（A20, A40, UPS, 管理机）
            m = re.match(r'^(A\d+|施耐德UPS|管理机)', col_a)
            if m:
                template_type = m.group(1)
                # 提取AI和DI数量
                ai_match = re.search(r'AI\s*(\d+)', col_a)
                di_match = re.search(r'DI\s*(\d+)', col_a)
                ai_count = int(ai_match.group(1)) if ai_match else 0
                di_count = int(di_match.group(1)) if di_match else 0
                
                current_template = {
                    'type': template_type,
                    'aiCount': ai_count,
                    'diCount': di_count,
                    'aiParams': [],
                    'diParams': []
                }
                templates.append(current_template)
                print(f"\n[模板类型] {template_type}: AI={ai_count}, DI={di_count}")
        
        # 如果当前有模板类型，收集AI和DI参数定义
        if current_template:
            # 收集AI参数（模板定义，只在模板类型行之后的几行）
            if ai_name and ai_name not in ['AI名称', '']:
                ai_param = {
                    'name': ai_name,
                    'offset': offset,
                    'factor': factor,
                    'parseMode': parse_mode
                }
                # 避免重复添加相同的参数
                if ai_param not in current_template['aiParams']:
                    current_template['aiParams'].append(ai_param)
            
            # 收集DI参数
            if di_name and di_name not in ['DI名称', '']:
                di_param = {
                    'name': di_name,
                    'registerOffset': di_reg_offset,
                    'bitOffset': di_bit_offset
                }
                if di_param not in current_template['diParams']:
                    current_template['diParams'].append(di_param)
        
        # 收集设备实例（只要有设备名称就记录）
        if device_name and device_name not in ['设备名称', '']:
            devices_from_template.append({
                'name': device_name,
                'templateType': current_template['type'] if current_template else '未知',
                'aiStartAddr': int(ai_start_addr) if isinstance(ai_start_addr, (int, float)) else ai_start_addr,
                'diStartAddr': int(di_start_addr) if isinstance(di_start_addr, (int, float)) else di_start_addr,
                'aiName': ai_name,
                'diName': di_name
            })
    
    print(f"\n[模板] 发现 {len(templates)} 种模板类型")
    for t in templates:
        print(f"  {t['type']}: AI参数={len(t['aiParams'])}, DI参数={len(t['diParams'])}")
    
    print(f"\n[模板] 发现 {len(devices_from_template)} 个设备实例")
    for d in devices_from_template[:10]:
        print(f"  {d['name']} (模板: {d['templateType']}, AI起始: {d['aiStartAddr']}, DI起始: {d['diStartAddr']})")
    if len(devices_from_template) > 10:
        print(f"  ... 还有 {len(devices_from_template) - 10} 个")
    
    # ========== 3. 读取Sheet1（设备列表）==========
    sheet1 = wb['Sheet1']
    s1_rows = list(sheet1.iter_rows(min_row=1, values_only=True))
    
    devices_sheet1 = []
    for i, row in enumerate(s1_rows):
        if i == 0:  # 跳过表头
            continue
        if row[0] and str(row[0]).strip():
            device_name = str(row[0]).strip()
            device_type = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            ai_count = row[2] if len(row) > 2 and row[2] is not None else 0
            di_count = row[3] if len(row) > 3 and row[3] is not None else 0
            
            devices_sheet1.append({
                'name': device_name,
                'type': device_type,
                'aiCount': int(ai_count) if isinstance(ai_count, (int, float)) else 0,
                'diCount': int(di_count) if isinstance(di_count, (int, float)) else 0
            })
    
    print(f"\n[Sheet1] 设备列表: {len(devices_sheet1)} 个设备")
    
    # ========== 4. 分析主表数据点结构 ==========
    print("\n=== 主表数据点结构分析 ===")
    
    # 按源节点分析子设备
    for src_name, points in sorted(device_points.items(), key=lambda x: -len(x[1])):
        print(f"\n{src_name} ({len(points)} 个数据点)")
        
        # 从遥信名提取子设备
        sub_devices = defaultdict(list)
        for p in points:
            yx = p['yx_name']
            # 匹配模式: "P1_仪表_ 1A1_U11_S18_1Input 1" -> 子设备 "1A1_U11_S18"
            # 或者 "P1_仪表_ 1A1_U11_S18_1U12_1" -> 子设备 "1A1_U11_S18_1"
            m = re.search(r'\s+([\w\-]+_S\d+)(?:_\d+)?', yx)
            if not m:
                m = re.search(r'\s+([\w\-]+_S\d+)', yx)
            
            sub_dev = m.group(1) if m else '其他'
            sub_devices[sub_dev].append(p)
        
        # 打印子设备统计
        for sub_dev, sub_pts in sorted(sub_devices.items(), key=lambda x: -len(x[1]))[:5]:
            print(f"  {sub_dev}: {len(sub_pts)} 个数据点")
            # 统计参数类型
            param_types = defaultdict(int)
            for p in sub_pts:
                # 匹配参数名如 U12, I1, F, P, Q 等
                pm = re.search(r'_(?:\d+)?([A-Za-z]+)_(?:\d+)?$', p['yx_name'])
                if pm:
                    param_types[pm.group(1)] += 1
                else:
                    pm2 = re.search(r'Input\s+(\d+)', p['yx_name'])
                    if pm2:
                        param_types['Input'] += 1
                    else:
                        param_types['其他'] += 1
            
            if param_types:
                print(f"    参数分布: {dict(param_types)}")
            # 只显示前3个示例
            for p in sub_pts[:3]:
                print(f"    - {p['yx_name'][:50]} (寄存器: {p['reg_addr']})")
    
    # ========== 5. 生成ISM数据模型 ==========
    print("\n\n=== 生成ISM数据模型 ===")
    
    models = {}
    
    # 遍历模板类型，生成模型
    for template in templates:
        model_name = template['type']
        
        # 构建AI数据点列表
        ai_points = []
        for ai in template['aiParams']:
            ai_points.append({
                'name': ai['name'],
                'dataType': 'FLOAT32',  # 根据解析方式判断，默认FLOAT32
                'registerType': 'HOLDING',
                'registerCount': 2,  # FLOAT32占2个寄存器
                'factor': ai['factor'],
                'parseMode': ai['parseMode']
            })
        
        # 构建DI数据点列表
        di_points = []
        for di in template['diParams']:
            di_points.append({
                'name': di['name'],
                'dataType': 'BOOL',
                'registerType': 'DI',
                'registerCount': 1,
                'registerOffset': di['registerOffset'],
                'bitOffset': di['bitOffset']
            })
        
        models[model_name] = {
            'name': model_name,
            'description': f'{model_name} 设备模型',
            'protocol': 'ModbusTCP',
            'aiCount': template['aiCount'],
            'diCount': template['diCount'],
            'aiPoints': ai_points,
            'diPoints': di_points
        }
    
    # 保存结果
    result = {
        'sourceNodes': {k: len(v) for k, v in device_points.items()},
        'templates': templates,
        'devicesFromTemplate': devices_from_template,
        'devicesFromSheet1': devices_sheet1,
        'models': models
    }
    
    output_path = Path(file_path).parent / 'ism_data_models.json'
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    
    print(f"\n✅ 数据模型已保存到: {output_path}")
    print(f"\n模型摘要:")
    for name, model in models.items():
        print(f"  {name}: AI={len(model['aiPoints'])}, DI={len(model['diPoints'])}")
    
    return result


if __name__ == '__main__':
    import sys
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
    else:
        file_path = '/Users/yunfanleo/cursorProjects/ISM源码/1A配电室 172.31.4.14 172.20.255.14.xlsx'
    
    analyze_excel(file_path)
