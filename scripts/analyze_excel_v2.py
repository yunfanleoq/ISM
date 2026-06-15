#!/usr/bin/env python3
"""
ISM Excel 数据点分析器 - 修正版
正确提取主表数据点，按设备类型归类，生成设备模型
"""
import openpyxl
import json
import re
from collections import defaultdict
from pathlib import Path


def analyze_excel(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # ========== 1. 读取主表（数据点列表）==========
    main_sheet = wb['1A配电室 172.31.4.14 172.20.255.14']
    all_rows = list(main_sheet.iter_rows(min_row=1, values_only=True))
    
    # 找到表头行
    header_row_idx = None
    for i, row in enumerate(all_rows):
        if row[0] and str(row[0]).strip() == '点号':
            header_row_idx = i
            break
    
    headers = [str(c).strip() if c else '' for c in all_rows[header_row_idx]]
    print(f"[主表] 表头: {headers}")
    
    # 读取数据行
    data_rows = []
    for i in range(header_row_idx + 1, len(all_rows)):
        row = all_rows[i]
        if row[0] is not None and str(row[0]).strip() != '' and str(row[0]).strip() != '点号':
            data_rows.append(row)
    
    print(f"[主表] 数据行数: {len(data_rows)}")
    
    # 按源节点分组
    device_points = defaultdict(list)
    for row in data_rows:
        src_name = str(row[3]).strip() if len(row) > 3 and row[3] else ''
        yx_name = str(row[4]).strip() if len(row) > 4 and row[4] else ''
        reg_addr = row[5] if len(row) > 5 and row[5] is not None else ''
        
        if src_name and src_name != '源节点名':
            device_points[src_name].append({
                'yx_name': yx_name,
                'reg_addr': int(reg_addr) if isinstance(reg_addr, (int, float)) else str(reg_addr)
            })
    
    print(f"[主表] 源节点数: {len(device_points)}")
    
    # ========== 2. 读取模板表（设备类型定义）==========
    template_sheet = wb['模板']
    tpl_rows = list(template_sheet.iter_rows(min_row=1, values_only=True))
    
    # 找到模板表表头
    tpl_header_idx = None
    for i, row in enumerate(tpl_rows):
        if row[0] and str(row[0]).strip() == '点位名称':
            tpl_header_idx = i
            break
    
    tpl_headers = [str(c).strip() if c else '' for c in tpl_rows[tpl_header_idx]]
    print(f"\n[模板] 表头: {tpl_headers}")
    
    # 读取模板数据，按设备类型分组
    templates = defaultdict(list)
    current_type = None
    for i in range(tpl_header_idx + 1, len(tpl_rows)):
        row = tpl_rows[i]
        if row[0] and str(row[0]).strip():
            val = str(row[0]).strip()
            # 检查是否是设备类型标题行
            if '电力仪表' in val or 'UPS' in val or '管理机' in val:
                current_type = val
                continue
            
            if current_type:
                point_name = str(row[0]).strip() if row[0] else ''
                data_type = str(row[1]).strip() if len(row) > 1 and row[1] else 'UINT16'
                reg_type = str(row[2]).strip() if len(row) > 2 and row[2] else 'HOLDING'
                
                templates[current_type].append({
                    'name': point_name,
                    'dataType': data_type,
                    'registerType': reg_type
                })
    
    print(f"[模板] 设备类型数: {len(templates)}")
    for t, pts in templates.items():
        print(f"  {t}: {len(pts)} 个数据点")
    
    # ========== 3. 读取Sheet1（设备列表）==========
    sheet1 = wb['Sheet1']
    s1_rows = list(sheet1.iter_rows(min_row=1, values_only=True))
    
    devices = []
    for i, row in enumerate(s1_rows):
        if i == 0:  # 跳过表头
            continue
        if row[0] and str(row[0]).strip():
            device_name = str(row[0]).strip()
            device_type = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            ai_count = row[2] if len(row) > 2 and row[2] is not None else 0
            di_count = row[3] if len(row) > 3 and row[3] is not None else 0
            
            devices.append({
                'name': device_name,
                'type': device_type,
                'aiCount': int(ai_count) if isinstance(ai_count, (int, float)) else 0,
                'diCount': int(di_count) if isinstance(di_count, (int, float)) else 0
            })
    
    # ========== 4. 分析每个源节点下的子设备结构 ==========
    print("\n=== 各源节点子设备结构 ===")
    for src_name, points in sorted(device_points.items(), key=lambda x: -len(x[1])):
        print(f"\n{src_name} ({len(points)} 个数据点)")
        
        # 从遥信名提取子设备
        sub_devices = defaultdict(list)
        for p in points:
            yx = p['yx_name']
            # 匹配模式: "P1_仪表_ 1A1_U11_S18_1Input 1" -> 子设备 "1A1_U11_S18_1"
            m = re.search(r'\s+([\w\-]+_S\d+)_\d+', yx)
            if not m:
                m = re.search(r'\s+([\w\-]+_S\d+)[\s_]', yx)
            
            sub_dev = m.group(1) if m else '其他'
            sub_devices[sub_dev].append(p)
        
        # 打印子设备统计
        for sub_dev, sub_pts in sorted(sub_devices.items(), key=lambda x: -len(x[1])):
            print(f"  {sub_dev}: {len(sub_pts)} 个数据点")
            # 统计参数类型
            param_types = defaultdict(int)
            for p in sub_pts:
                pm = re.search(r'_([A-Za-z]+)_\d+', p['yx_name'])
                if pm:
                    param_types[pm.group(1)] += 1
                else:
                    pm2 = re.search(r'Input\s+(\d+)', p['yx_name'])
                    if pm2:
                        param_types['Input'] += 1
                    else:
                        param_types['其他'] += 1
            
            if param_types:
                print(f"    参数分布: {dict(param_types)}")
            # 只显示前3个示例
            for p in sub_pts[:3]:
                print(f"    - {p['yx_name'][:50]} (寄存器: {p['reg_addr']})")
    
    # ========== 5. 生成设备模型 ==========
    print("\n\n=== 生成设备模型 ===")
    
    # 分析每个源节点对应的设备类型
    models = {}
    for src_name, points in device_points.items():
        # 判断设备类型
        if 'UPS' in src_name:
            model_type = '施耐德UPS'
        elif '仪表' in src_name:
            # 根据数据点数量判断是A20还是A40
            if len(points) >= 300:  # P1/P2/P4/P5/P7 有 306 个数据点
                model_type = 'A20电力仪表'
            else:
                model_type = 'A40电力仪表'
        else:
            model_type = '管理机'
        
        # 提取该源节点的数据点结构
        model_points = []
        
        # 分析参数结构
        for p in points[:50]:  # 取前50个分析结构
            yx = p['yx_name']
            reg = p['reg_addr']
            
            # 判断是AI还是DI
            if 'Input' in yx or '通讯状态' in yx or '运行状态' in yx:
                point_type = 'DI'
            else:
                point_type = 'AI'
            
            # 提取参数名
            param_match = re.search(r'_([A-Za-z]+)_\d+$', yx)
            if param_match:
                param_name = param_match.group(1)
            else:
                param_name = '未知'
            
            model_points.append({
                'name': yx,
                'type': point_type,
                'param': param_name,
                'register': reg
            })
        
        models[src_name] = {
            'type': model_type,
            'totalPoints': len(points),
            'samplePoints': model_points
        }
    
    # 保存结果
    result = {
        'sourceNodes': {k: len(v) for k, v in device_points.items()},
        'templates': {k: len(v) for k, v in templates.items()},
        'deviceList': devices,
        'models': models
    }
    
    output_path = Path(file_path).parent / 'analysis_result.json'
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    
    print(f"\n分析结果已保存到: {output_path}")
    return result


if __name__ == '__main__':
    import sys
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
    else:
        file_path = '/Users/yunfanleo/cursorProjects/ISM源码/1A配电室 172.31.4.14 172.20.255.14.xlsx'
    
    analyze_excel(file_path)
