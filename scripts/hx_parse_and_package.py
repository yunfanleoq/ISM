#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""中航信数据中心 —— 转发表解析 + ISM 项目包生成（三格式统一适配）。

输入: 数据点位转发表/_unzipped/{A,模块,配电室}/*.xlsx
输出(每个数据集一套):
  <outdir>/ism_data_models.json              —— devices[] + models{}（建项目/建设备用）
  <outdir>/中航信_complete_project_package.json —— deviceModels/registerGroups/registerPoints（建数据点用）
  <outdir>/parse_report.json                 —— 解析统计（人工核对）

三种 Excel 模式:
  A  (excel(1).zip)  —— 有 '模板' sheet，多模型(A20/A40/施耐德UPS/伊顿UPS)
  模块(20260613.zip) —— 有 '汇总' sheet，单模型(列头柜)，SJJF* 命名
  配电室(TB.zip)     —— 有 'Sheet1'，单模型(仪表)，*_DX* 命名

用法:
  python3 scripts/hx_parse_and_package.py --set validation
  python3 scripts/hx_parse_and_package.py --set full
"""
import os, re, json, glob, argparse, hashlib, uuid
from collections import defaultdict, OrderedDict
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
UNZIP = os.path.join(ROOT, "数据点位转发表", "_unzipped")

# 验证集: 三种格式各一个典型文件
VALIDATION = {
    "A":   ["1B配电室 172.31.4.15 172.20.255.15.xlsx"],
    "模块": ["2A3模块172.20.255.25.xlsx"],
    "配电室": ["2A2配电室172.31.4.24.xlsx"],
}

PROJECT_NAME = "中航信数据中心电力监控系统"

# parse_mode → (Go Type, 寄存器数)。我们同时用模拟器自产数据，类型只要前后端一致即可。
PARSE_MODE_TYPE = {
    73:  ("Short", 1), 177: ("Unsigned short", 1), 71: ("Unsigned short", 1),
    1:   ("Unsigned short", 1),
    179: ("Long", 2), 133: ("Long", 2), 171: ("Long", 2),
}


def gen_uuid():
    return uuid.uuid4().hex


def infer_type(parse_mode, span, name=""):
    """推断 Go 类型与寄存器数。

    寄存器宽度以「偏移跨度 span」为物理铁证(转发表「寄存器偏移」列即真实寄存器号):
      - span==1  → 必为 1 寄存器。即便 parse_mode 标成 Long(如列头柜主回路/支路有功功率
        parse=133,但偏移 16,17,18 连续),该设备就只占 1 个寄存器;若按 2 寄存器解析,相邻点
        会重叠,后端「顺序消费寄存器」会吞掉一半点位(实测列头柜 40点/台空值的根因)。
      - span>=2  → 宽度由 parse_mode 决定: A 系列电压(parse=单字)按 2 间隔但只占 1 寄存器
        (奇地址留空); 电度(parse=Long)且偏移跨度=2 → 真 2 寄存器。
      - span 缺失(分段末点) → 信任 parse_mode。
    类型: 2 寄存器统一按 Float(CDAB) 处理(模拟器与后端自洽即可); 1 寄存器按 parse 有/无符号,
          由 Long 降级而来的取 Short(允许反向负功率), 畸变率取 Short。
    """
    try:
        pm = int(float(parse_mode))
    except (TypeError, ValueError):
        pm = None
    base = PARSE_MODE_TYPE.get(pm)            # (type, width) 或 None
    base_type, base_w = base if base else (None, None)

    def one_reg():
        if base_w == 1:
            return base_type
        if base_w == 2 or "畸变率" in name:
            return "Short"          # Long 降级 / 畸变率 → 有符号 16 位
        return "Unsigned short"

    def two_or_one():
        if base_w:
            return ("Float", 2) if base_w >= 2 else (base_type, 1)
        if "畸变率" in name:
            return ("Short", 1)
        return ("Float", 2)         # 未知 parse 且跨度>=2 → 视为 2 寄存器

    if span == 1:
        return (one_reg(), 1)
    if span is None:
        if base_w:
            return ("Float", 2) if base_w >= 2 else (base_type, 1)
        return ("Short", 1) if "畸变率" in name else ("Unsigned short", 1)
    return two_or_one()              # span >= 2


def s(v):
    return str(v).strip() if v is not None else ""


def num(v):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def detect_format(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    if "模板" in names:
        return "A"
    if "汇总" in names:
        return "模块"
    if "Sheet1" in names and "两字节遥测" in names:
        return "配电室"
    return None


# ---------- 偏移跨度推断（为一组有序 AI 点计算寄存器数）----------
def assign_spans(ai_points):
    offs = [p["offset"] for p in ai_points]
    for i, p in enumerate(ai_points):
        nxt = None
        for j in range(i + 1, len(ai_points)):
            if offs[j] is not None and offs[i] is not None and offs[j] > offs[i]:
                nxt = offs[j]; break
        p["span"] = (nxt - p["offset"]) if (nxt is not None and p["offset"] is not None) else None


# =================== 统一“分段式”解析器 ===================
KNOWN_PARSE = {177, 179, 73, 71, 1, 133, 171, 70}
HEADERS = {"AI名称", "DI名称", "设备名称", "组态名称", "组态命名", "名称替换",
           "模板类型", "模版类型", "寄存器偏移", "处理方式", "解析方式", "系数"}


def _is_factor(v):
    try:
        f = float(v)
        return 0 < f < 1 or f in {0.1, 0.01, 0.001, 0.008, 0.02, 0.08}
    except (TypeError, ValueError):
        return False


def _is_parse(v):
    try:
        f = float(v)
        return f in KNOWN_PARSE or f > 100
    except (TypeError, ValueError):
        return False


def _factor_pm(a, b):
    """从两列里区分 (系数, 解析方式)，自动处理交换。"""
    if _is_factor(a) and _is_parse(b):
        return a, b
    if _is_factor(b) and _is_parse(a):
        return b, a
    return a, b


def _clean_marker(v):
    first = str(v).split("\n")[0]
    first = re.sub(r"\s*\d+\s*个.*$", "", first)
    return first.strip()


# =================== 表头自动映射（取代硬编码列号）===================
# 全部 56 个文件、16 种签名共用同一套表头词汇，仅列位置漂移。
# 不再按文件夹硬编码列号——按表头关键词自动定位每个逻辑字段，一招覆盖所有变体。
HDR_KW = {"AI名称", "DI名称", "模板类型", "模版类型", "设备名称", "组态命名", "组态名称",
          "名称替换", "寄存器偏移", "偏移", "处理方式", "解析方式", "位偏移", "系数"}
OFFSET_HDR = {"寄存器偏移", "偏移"}     # AI/DI 偏移列(A 系列用“偏移”,其余用“寄存器偏移”); 不含“位偏移”


def _header_score(cells):
    return sum(1 for c in cells if c in HDR_KW)


def _find_header(rows, max_scan=8):
    """返回命中关键词最多的表头行下标(>=3 命中才算)。"""
    best = (2, -1)
    for i, r in enumerate(rows[:max_scan]):
        sc = _header_score([s(c) for c in r])
        if sc > best[0]:
            best = (sc, i)
    return best[1]


def _build_colmap(header_cells):
    """按表头关键词构造逻辑字段→列号映射。"""
    def idx(*names):
        return next((i for i, c in enumerate(header_cells) if c in names), None)

    def off_idx():
        return [i for i, c in enumerate(header_cells) if c in OFFSET_HDR]

    cm = {
        "marker":   idx("模板类型", "模版类型"),
        "ai_name":  idx("AI名称"),
        "di_name":  idx("DI名称"),
        "ai_parse": idx("处理方式", "解析方式"),
        "ai_factor": idx("系数"),
        "di_bit":   idx("位偏移"),
        "dev_main": idx("组态命名", "组态名称"),   # 组态名优先作为设备 SN
        "dev_alt":  idx("设备名称"),
        "disp":     idx("名称替换"),
    }
    offs = off_idx()
    cm["ai_off"] = next((o for o in offs if cm["ai_name"] is not None and o > cm["ai_name"]), None)
    cm["di_off"] = next((o for o in offs if cm["di_name"] is not None and o > cm["di_name"]), None)
    return cm


def _sheet_ok(ws):
    rows = list(ws.iter_rows(min_row=1, max_row=8, values_only=True))
    hi = _find_header(rows)
    if hi < 0:
        return False
    cm = _build_colmap([s(c) for c in rows[hi]])
    return cm["marker"] is not None and cm["ai_name"] is not None


def _pick_sheet(wb):
    """选数据 sheet：汇总 > 模板 > Sheet1 > 表头命中最多者。"""
    names = wb.sheetnames
    for pref in ("汇总", "模板"):
        if pref in names and _sheet_ok(wb[pref]):
            return pref
    if "Sheet1" in names and _sheet_ok(wb["Sheet1"]):
        return "Sheet1"
    best = (2, None)
    for sn in names:
        try:
            sc = _header_score([s(c) for c in next(wb[sn].iter_rows(min_row=1, max_row=1, values_only=True))])
        except StopIteration:
            sc = 0
        if sc > best[0] and _sheet_ok(wb[sn]):
            best = (sc, sn)
    return best[1]


def parse_sectioned(path, fmt=None):
    """统一解析：自动选 sheet + 自动映射列，按 marker 切分模型段。

    返回 sections = [{type, ai:[], di:[], devices:[{sn, display}]}]
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = _pick_sheet(wb)
    if sheet is None:
        wb.close()
        return []
    rows = list(wb[sheet].iter_rows(values_only=True))
    wb.close()
    hi = _find_header(rows)
    if hi < 0:
        return []
    cm = _build_colmap([s(c) for c in rows[hi]])
    if cm["marker"] is None or cm["ai_name"] is None:
        return []

    def cell(row, idx):
        return row[idx] if idx is not None and len(row) > idx else None

    sections = []
    cur = None
    for row in rows[hi + 1:]:
        mk = s(cell(row, cm["marker"]))
        if mk and mk not in HEADERS:
            cur = {"type": _clean_marker(mk), "ai": [], "di": [], "devices": []}
            sections.append(cur)
        if cur is None:
            continue
        ai_name = s(cell(row, cm["ai_name"]))
        if ai_name and ai_name not in HEADERS:
            factor, pm = _factor_pm(cell(row, cm["ai_factor"]), cell(row, cm["ai_parse"]))
            p = {"name": ai_name, "offset": num(cell(row, cm["ai_off"])),
                 "factor": factor, "parseMode": pm}
            if p not in cur["ai"]:
                cur["ai"].append(p)
        di_name = s(cell(row, cm["di_name"])) if cm["di_name"] is not None else ""
        if di_name and di_name not in HEADERS:
            p = {"name": di_name, "offset": num(cell(row, cm["di_off"])),
                 "bit": num(cell(row, cm["di_bit"]))}
            if p not in cur["di"]:
                cur["di"].append(p)
        sn = s(cell(row, cm["dev_main"])) or s(cell(row, cm["dev_alt"]))
        if sn and sn not in HEADERS:
            disp = s(cell(row, cm["disp"])) or sn
            cur["devices"].append({"sn": sn, "display": disp})
    for sec in sections:
        assign_spans(sec["ai"])
    return sections


# =================== 文件名 → 楼栋/室/IP ===================
def parse_filename(fname):
    base = os.path.basename(fname)
    ips = re.findall(r"\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}", base)
    primary = next((ip for ip in ips if ip.startswith("172.31")), ips[0] if ips else "127.0.0.1")
    backup = next((ip for ip in ips if ip.startswith("172.20")), "")
    head = re.split(r"\s|\d{1,3}\.\d{1,3}\.", base)[0]
    head = re.sub(r"\.xlsx?$", "", head).strip()
    mb = re.match(r"(ECC\d?|\d[AB]\d?)", head)
    building = ""
    if mb:
        token = mb.group(1)
        building = token[:3] if token.startswith("ECC") else token[:2]
    room = head or base
    return building, room, primary, backup


# =================== 楼栋/室/柜 层级 ===================
def hierarchy_for(sn, building, room, fmt):
    # 注意: 柜级区域名必须与设备 SN 不冲突(否则同名 monitor_list 互相覆盖)，统一加“柜”后缀
    parts = sn.split("_")
    if fmt == "A":
        # 1B1_U11_S14_1 / 1B1_U11(UPS两段)
        sub = parts[0] if parts else building
        cab = ("_".join(parts[:2]) if len(parts) >= 2 else sub) + "柜"
        return building, room, cab
    if fmt == "配电室":
        # 2A1_U2_DX17_1
        cab = (parts[1] if len(parts) >= 2 else room) + "柜"
        return building, room, cab
    if fmt == "模块":
        # SJJF2A3_1A -> 列分组
        m = re.search(r"_(\d+)([A-Z])$", sn)
        cab = f"{room}-{m.group(1)}列" if m else (room + "-列")
        return building, room, cab
    return building, room, room + "-柜"


# =================== 模型去重（按布局 hash）===================
class ModelRegistry:
    def __init__(self):
        self.by_hash = {}     # hash -> model dict
        self.name_count = defaultdict(int)
        self.models = []      # 输出顺序

    def _hash(self, mt, spec):
        key = json.dumps([mt, [(p["name"], p["offset"], p.get("span")) for p in spec["ai"]],
                          [(p["name"], p["offset"]) for p in spec["di"]], ], ensure_ascii=False)
        return hashlib.md5(key.encode()).hexdigest()

    def get_or_create(self, mt, spec):
        h = self._hash(mt, spec)
        if h in self.by_hash:
            return self.by_hash[h]
        self.name_count[mt] += 1
        name = mt if self.name_count[mt] == 1 else f"{mt}_{self.name_count[mt]}"
        muid = gen_uuid()
        ai_uuid, di_uuid = gen_uuid(), gen_uuid()
        ai_count = (max([p["offset"] for p in spec["ai"] if p["offset"] is not None] + [0]) + 2) if spec["ai"] else 1
        di_count = (max([p["offset"] for p in spec["di"] if p["offset"] is not None] + [0]) + 1) if spec["di"] else 1
        model = {
            "name": name, "modelType": mt, "muid": muid,
            "aiGroupUuid": ai_uuid, "diGroupUuid": di_uuid,
            "aiCount": ai_count, "diCount": di_count,
            "ai": spec["ai"], "di": spec["di"],
        }
        self.by_hash[h] = model
        self.models.append(model)
        return model


def build(dataset_files, outdir):
    os.makedirs(outdir, exist_ok=True)
    reg = ModelRegistry()
    devices_out = []        # for ism_data_models.json
    report = {"files": [], "totalDevices": 0}
    sn_seen = set()         # 去重键 (building, room, sn)：同室同名才算重复
    name_used = set()       # 全局唯一设备名集合

    for fmt, fname in dataset_files:
        path = os.path.join(UNZIP, fmt, fname)
        if not os.path.exists(path):
            print(f"  !! 缺文件: {path}"); continue
        building, room, ip, backup = parse_filename(fname)
        # 解析器自动选 sheet + 自动映射列；解析不出任何段则跳过(空表/无表头)。
        try:
            sections = parse_sectioned(path)
        except Exception as e:
            print(f"  -- 跳过(解析失败 {type(e).__name__}): {fname}")
            continue
        if not sections:
            print(f"  -- 跳过(无可识别表头/数据): {fname}")
            continue
        # 逐段注册模型并归属设备
        n_file = 0
        file_models = []
        for sec in sections:
            if not sec["ai"] and not sec["di"]:
                continue
            model = reg.get_or_create(sec["type"], sec)
            file_models.append(model["name"])
            for d in sec["devices"]:
                sn = d["sn"]
                key = (building, room, sn)
                if key in sn_seen:       # 仅同楼栋同室同名才算重复
                    continue
                sn_seen.add(key)
                # 生成全局唯一设备名：通用短名(如 P1_TA)在不同楼栋会重名，按 室/楼栋 命名空间消歧
                uname = sn
                if uname in name_used:
                    uname = f"{room}_{sn}"
                if uname in name_used:
                    uname = f"{building}_{room}_{sn}"
                while uname in name_used:
                    uname += "_x"
                name_used.add(uname)
                b, r, cab = hierarchy_for(sn, building, room, fmt)
                devices_out.append({
                    "name": uname, "display": d.get("display", sn),
                    "modelName": model["name"], "templateType": sec["type"],
                    "building": b, "room": r, "cabinet": cab,
                    "gatewayIP": ip, "gatewayBackup": backup, "port": 502,
                })
                n_file += 1
        report["files"].append({"file": fname, "fmt": fmt, "building": building,
                                "room": room, "ip": ip, "devices": n_file,
                                "models": file_models})
        print(f"  [{fmt}] {fname}: 楼栋={building} 室={room} IP={ip} 设备={n_file} 模型={file_models}")

    report["totalDevices"] = len(devices_out)

    # ---- 分配 slave id（按网关 IP 内 1..n；Modbus 从站号上限 247，全局唯一会溢出）----
    gw_counter = defaultdict(int)
    for d in devices_out:
        gw_counter[d["gatewayIP"]] += 1
        d["slaveId"] = gw_counter[d["gatewayIP"]]

    # ---- 生成两份契约 ----
    # 1) ism_data_models.json
    models_json = {
        "project": PROJECT_NAME,
        "devices": devices_out,
        "models": {m["name"]: {
            "name": m["name"], "modelType": m["modelType"],
            "aiCount": m["aiCount"], "diCount": m["diCount"],
            "aiPoints": m["ai"], "diPoints": m["di"],
        } for m in reg.models},
    }
    with open(os.path.join(outdir, "ism_data_models.json"), "w", encoding="utf-8") as f:
        json.dump(models_json, f, ensure_ascii=False, indent=2)

    # 2) complete_project_package.json
    proj_uuid = gen_uuid()
    MAX_GRP = 100   # 单次 Modbus 读 <=125 寄存器，留余量按 100 切组
    deviceModels, registerGroups, registerPoints = [], [], []

    def ai_group_for(offset, ai_chunks):
        """返回覆盖该 offset 的 AI 组 uuid。"""
        for (lo, hi, guid) in ai_chunks:
            if lo <= offset < hi:
                return guid
        return ai_chunks[-1][2]

    for m in reg.models:
        deviceModels.append({
            "uuid": m["muid"], "name": m["name"], "dec": f"{m['name']} 设备模型",
            "type": 2, "gatherNumber": 30, "project_uuid": proj_uuid,
            # DataFormat=ABCD: 单寄存器(Short/Unsigned short)按大端解析，与模拟器 '>H' 直接匹配；
            # 多寄存器点用各自 ByteOrder=CDAB(Float) 独立控制，不受 DataFormat 影响。
            "DataFormat": "ABCD", "modbusConnectType": "TCPClient",
            "modbusConnectMode": "TCP/IP",
        })
        # AI 组按 <=100 寄存器切分(覆盖大模型如列头柜 478 寄存器)
        ai_chunks = []
        n_chunks = max(1, (m["aiCount"] + MAX_GRP - 1) // MAX_GRP)
        for ci in range(n_chunks):
            lo = ci * MAX_GRP
            cnt = min(MAX_GRP, m["aiCount"] - lo)
            guid = m["aiGroupUuid"] if ci == 0 else gen_uuid()
            gname = "AI数据" if n_chunks == 1 else f"AI数据_{ci+1}"
            registerGroups.append({"uuid": guid, "muid": m["muid"], "name": gname,
                                   "function": 3, "registerStart": lo, "registerCount": cnt})
            ai_chunks.append((lo, lo + cnt, guid))
        registerGroups.append({"uuid": m["diGroupUuid"], "muid": m["muid"], "name": "DI数据",
                               "function": 2, "registerStart": 0, "registerCount": m["diCount"]})
        for p in m["ai"]:
            if p["offset"] is None:
                continue
            gt, rc = infer_type(p.get("parseMode"), p.get("span"), p["name"])
            try:
                fac = float(p.get("factor")) if p.get("factor") not in (None, "") else 1.0
            except (TypeError, ValueError):
                fac = 1.0
            if rc >= 2:
                # 双寄存器统一用 Float(CDAB)：寄存器直接承载物理真值，无需换算
                gt = "Float"
                conv = ""
                sim_factor = 1.0       # 模拟器直接写真值
            else:
                # 单寄存器整型：寄存器=真值/系数，前端用 {val}*系数 还原
                conv = "" if fac == 1.0 else f"{{val}}*{fac:g}"
                sim_factor = fac
            registerPoints.append({
                "uuid": gen_uuid(), "muid": m["muid"], "name": p["name"],
                "registerAddress": p["offset"], "registerGroupUuid": ai_group_for(p["offset"], ai_chunks),
                "auth": "ReadOnly", "type": gt, "ByteOrder": "CDAB", "modeltype": 2,
                "unit": guess_unit(p["name"]), "conversionExpression": conv,
                "factor": fac, "simFactor": sim_factor, "registerCount": rc,
                "record": 0, "RecordType": 1, "RecordInterval": 5,
                "FloatAccuracy": (f"{fac:g}" if fac != 1.0 else "1"),
            })
        for p in m["di"]:
            if p["offset"] is None:
                continue
            registerPoints.append({
                "uuid": gen_uuid(), "muid": m["muid"], "name": p["name"],
                "registerAddress": p["offset"], "registerGroupUuid": m["diGroupUuid"],
                "auth": "ReadOnly", "type": "Bool", "ByteOrder": "CDAB", "modeltype": 2,
                "unit": "", "conversionExpression": "", "factor": 1.0, "registerCount": 1,
                "record": 0, "RecordType": 1, "RecordInterval": 5, "FloatAccuracy": "1",
            })
    pkg = {
        "project": {"uuid": proj_uuid, "name": PROJECT_NAME,
                    "description": "中航信数据中心电力监控系统", "protocol": "ModbusTCP"},
        "deviceModels": deviceModels,
        "registerGroups": registerGroups,
        "registerPoints": registerPoints,
        "statistics": {
            "totalDevices": len(devices_out),
            "deviceModelsCount": len(deviceModels),
            "registerGroupsCount": len(registerGroups),
            "registerPointsCount": len(registerPoints),
        },
    }
    with open(os.path.join(outdir, "中航信_complete_project_package.json"), "w", encoding="utf-8") as f:
        json.dump(pkg, f, ensure_ascii=False, indent=2)

    report["models"] = [{"name": m["name"], "type": m["modelType"],
                         "ai": len(m["ai"]), "di": len(m["di"]),
                         "aiCount": m["aiCount"], "diCount": m["diCount"]} for m in reg.models]
    with open(os.path.join(outdir, "parse_report.json"), "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)

    print(f"\n[汇总] 设备={len(devices_out)}  模型={len(deviceModels)}  寄存器组={len(registerGroups)}  数据点={len(registerPoints)}")
    print(f"[输出] {outdir}")
    return report


UNIT_RULES = [
    ("线电压", "V"), ("相电压", "V"), ("电压", "V"), ("电流", "A"), ("频率", "Hz"),
    ("有功功率", "kW"), ("无功功率", "kvar"), ("视在功率", "kVA"), ("功率因数", ""),
    ("有功", "kW"), ("无功", "kvar"), ("电度", "kWh"), ("电能", "kWh"),
    ("畸变率", "%"), ("温度", "℃"), ("湿度", "%"),
]


def guess_unit(name):
    for kw, u in UNIT_RULES:
        if kw in name:
            return u
    return ""


def collect_full():
    files = []
    for fmt in ("A", "模块", "配电室"):
        for p in sorted(glob.glob(os.path.join(UNZIP, fmt, "*.xlsx"))):
            base = os.path.basename(p)
            if base.startswith("~$") or base == "综合.xlsx":
                continue
            files.append((fmt, base))
    return files


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--set", choices=["validation", "full"], default="validation")
    args = ap.parse_args()
    if args.set == "validation":
        ds = [(fmt, fn) for fmt, lst in VALIDATION.items() for fn in lst]
        outdir = os.path.join(ROOT, "hx-data", "validation")
    else:
        ds = collect_full()
        outdir = os.path.join(ROOT, "hx-data", "full")
    print(f"=== 数据集: {args.set}  文件数={len(ds)} ===")
    build(ds, outdir)
