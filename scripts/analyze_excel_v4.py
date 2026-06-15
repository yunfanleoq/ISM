#!/usr/bin/env python3
"""
ISM Excel 数据点分析器 - 修正版 v4
正确处理模板表中系数和解析方式被交换的问题（A40/UPS）
"""
import openpyxl
import json
import re
from collections import defaultdict
from pathlib import Path


# parse_mode → (Go Type, register_count) 映射表
# 对应 Go 后端 modbusPthread.go 中 data.Type 的分支逻辑
PARSE_MODE_TYPE_MAP = {
    73:  ('Short', 1),           # int16, 1寄存器 → 谐波畸变率
    177: ('Unsigned short', 1),  # uint16, 1寄存器 → 电压/电流/频率
    179: ('Long', 2),            # int32, 2寄存器 → 功率/功率因数
    71:  ('Unsigned short', 1),  # uint16, 1寄存器 → UPS参数
}

def infer_data_type(parse_mode, data_name=''):
    """
    根据 parse_mode 推断 Go 后端的 data.Type 和 registerCount。
    返回 (go_type, register_count) 元组。
    
    Go 后端 modbusPthread.go 按 data.Type 分派解析分支:
      - "Short" → 1寄存器, int16
      - "Unsigned short" → 1寄存器, uint16
      - "Long" → 2寄存器, int32 (BigEndian/LittleEndian)
      - "Float" → 2寄存器, Float32 (CDAB/ABCD等字节序)
    
    ⚠️ 坑点: 之前硬编码所有点为 FLOAT32/2寄存器，导致 parse_mode=73 (int16) 
    的点被当作 Float 解析 → 读到垃圾值或 NaN → 前端显示为空。
    """
    try:
        pm = int(parse_mode)
    except (ValueError, TypeError):
        return ('Float', 2)  # 默认
    
    if pm in PARSE_MODE_TYPE_MAP:
        return PARSE_MODE_TYPE_MAP[pm]
    
    # 根据数据点名辅助判断
    if '畸变率' in data_name:
        return ('Short', 1)
    
    return ('Float', 2)  # 默认


def analyze_excel(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # ========== 1. 读取模板表（模板类型+设备实例）==========
    template_sheet = wb['模板']
    tpl_rows = list(template_sheet.iter_rows(min_row=1, values_only=True))
    
    print("=== 模板表结构分析 ===")
    
    templates = []
    current_template = None
    devices_from_template = []
    
    for i, row in enumerate(tpl_rows[1:], 1):  # 跳过表头
        col_a = str(row[0]).strip() if row[0] else ''
        ai_name = str(row[1]).strip() if len(row) > 1 and row[1] else ''
        offset = row[2] if len(row) > 2 and row[2] is not None else ''
        col_d = row[3] if len(row) > 3 and row[3] is not None else ''
        col_e = row[4] if len(row) > 4 and row[4] is not None else ''
        di_name = str(row[8]).strip() if len(row) > 8 and row[8] else ''
        di_reg_offset = row[9] if len(row) > 9 and row[9] is not None else ''
        di_bit_offset = row[10] if len(row) > 10 and row[10] is not None else ''
        device_name = str(row[11]).strip() if len(row) > 11 and row[11] else ''
        ai_start_addr = row[13] if len(row) > 13 and row[13] is not None else ''
        di_start_addr = row[14] if len(row) > 14 and row[14] is not None else ''
        
        # 检测模板类型行（包含 "AI" 和 "DI" 说明）
        if 'AI' in col_a and 'DI' in col_a and ('寄存器' in col_a or '个' in col_a):
            m = re.match(r'^(A\d+|施耐德UPS|管理机)', col_a)
            if m:
                template_type = m.group(1)
                ai_match = re.search(r'AI\s*(\d+)', col_a)
                di_match = re.search(r'DI\s*(\d+)', col_a)
                ai_count = int(ai_match.group(1)) if ai_match else 0
                di_count = int(di_match.group(1)) if di_match else 0
                
                current_template = {
                    'type': template_type,
                    'aiCount': ai_count,
                    'diCount': di_count,
                    'aiParams': [],
                    'diParams': []
                }
                templates.append(current_template)
                print(f"\n[模板类型] {template_type}: AI={ai_count}, DI={di_count}")
        
        # 如果当前有模板类型，收集AI和DI参数定义
        if current_template:
            # 判断系数和解析方式是否被交换
            # A20: 系数是小数(0.01, 0.001等)，解析方式是整数(177, 179, 73)
            # A40/UPS: 系数和解析方式被交换了，需要修正
            
            # 判断col_d和col_e哪个是系数，哪个是解析方式
            # 系数通常是小数或1，解析方式是整数（177, 179, 73, 71, 1等）
            KNOWN_PARSE = {177, 179, 73, 71, 1}  # 已知的解析方式码 (包括UPS code_1)
            def is_factor(val):
                if val == '' or val is None:
                    return False
                try:
                    f = float(val)
                    # 系数：小数(<1) 或 very small decimal, 排除已知解析码
                    return (f < 1 and f > 0) or f in {0.1, 0.01, 0.001}
                except:
                    return False
            
            def is_parse_mode(val):
                if val == '' or val is None:
                    return False
                try:
                    f = float(val)
                    return f in KNOWN_PARSE or f > 100  # 已知解析码 或 大整数
                except:
                    return False
            
            # 判断是否需要交换
            if is_factor(col_d) and is_parse_mode(col_e):
                # 正常格式：col_d=系数, col_e=解析方式
                factor = col_d
                parse_mode = col_e
            elif is_factor(col_e) and is_parse_mode(col_d):
                # 交换格式：col_d=解析方式, col_e=系数
                factor = col_e
                parse_mode = col_d
                print(f"  [修正] 交换系数和解析方式: {ai_name} (模板: {current_template['type']})")
            else:
                # 无法判断，使用原始值
                factor = col_d
                parse_mode = col_e
            
            # 收集AI参数（模板定义）
            if ai_name and ai_name not in ['AI名称', '']:
                ai_param = {
                    'name': ai_name,
                    'offset': offset,
                    'factor': factor,
                    'parseMode': parse_mode
                }
                # 避免重复添加相同的参数
                if ai_param not in current_template['aiParams']:
                    current_template['aiParams'].append(ai_param)
            
            # 收集DI参数
            if di_name and di_name not in ['DI名称', '']:
                di_param = {
                    'name': di_name,
                    'registerOffset': di_reg_offset,
                    'bitOffset': di_bit_offset
                }
                if di_param not in current_template['diParams']:
                    current_template['diParams'].append(di_param)
        
        # 收集设备实例（只要有设备名称就记录）
        if device_name and device_name not in ['设备名称', '']:
            devices_from_template.append({
                'name': device_name,
                'templateType': current_template['type'] if current_template else '未知',
                'aiStartAddr': int(ai_start_addr) if isinstance(ai_start_addr, (int, float)) else ai_start_addr,
                'diStartAddr': int(di_start_addr) if isinstance(di_start_addr, (int, float)) else di_start_addr,
                'aiName': ai_name,
                'diName': di_name
            })
    
    print(f"\n[模板] 发现 {len(templates)} 种模板类型")
    for t in templates:
        print(f"  {t['type']}: AI参数={len(t['aiParams'])}, DI参数={len(t['diParams'])}")
        # 打印前5个AI参数示例
        for ai in t['aiParams'][:5]:
            print(f"    AI: {ai['name']}, 偏移={ai['offset']}, 系数={ai['factor']}, 解析方式={ai['parseMode']}")
    
    print(f"\n[模板] 发现 {len(devices_from_template)} 个设备实例")
    
    # 按模板类型统计设备实例
    device_count_by_type = defaultdict(int)
    for d in devices_from_template:
        device_count_by_type[d['templateType']] += 1
    for t, count in device_count_by_type.items():
        print(f"  {t}: {count} 个设备")
    
    # ========== 2. 读取主表（数据点列表）==========
    main_sheet = wb['1A配电室 172.31.4.14 172.20.255.14']
    all_rows = list(main_sheet.iter_rows(min_row=1, values_only=True))
    
    # 找到表头行
    header_row_idx = None
    for i, row in enumerate(all_rows):
        if row[0] and str(row[0]).strip() == '点号':
            header_row_idx = i
            break
    
    # 读取数据行
    data_rows = []
    for i in range(header_row_idx + 1, len(all_rows)):
        row = all_rows[i]
        if row[0] is not None and str(row[0]).strip() != '' and str(row[0]).strip() != '点号':
            data_rows.append(row)
    
    # 按源节点分组
    device_points = defaultdict(list)
    for row in data_rows:
        src_name = str(row[3]).strip() if len(row) > 3 and row[3] else ''
        yx_name = str(row[4]).strip() if len(row) > 4 and row[4] else ''
        reg_addr = row[5] if len(row) > 5 and row[5] is not None else ''
        
        if src_name and src_name != '源节点名':
            device_points[src_name].append({
                'yx_name': yx_name,
                'reg_addr': int(reg_addr) if isinstance(reg_addr, (int, float)) else str(reg_addr)
            })
    
    print(f"\n[主表] 数据行数: {len(data_rows)}, 源节点数: {len(device_points)}")
    
    # ========== 3. 生成ISM数据模型 ==========
    print("\n=== 生成ISM数据模型 ===")
    
    models = {}
    for template in templates:
        model_name = template['type']
        
        # 构建AI数据点列表
        ai_points = []
        for ai in template['aiParams']:
            go_type, reg_count = infer_data_type(ai.get('parseMode', ''), ai.get('name', ''))
            ai_points.append({
                'name': ai['name'],
                'dataType': go_type,       # Go后端Type字段: Short/Unsigned short/Long/Float
                'registerType': 'HOLDING',
                'registerCount': reg_count, # 1寄存器(int16/uint16) 或 2寄存器(int32/Float32)
                'factor': ai['factor'],
                'parseMode': ai['parseMode'],
                'offset': ai['offset']
            })
        
        # 构建DI数据点列表
        di_points = []
        for di in template['diParams']:
            di_points.append({
                'name': di['name'],
                'dataType': 'BOOL',
                'registerType': 'DI',
                'registerCount': 1,
                'registerOffset': di['registerOffset'],
                'bitOffset': di['bitOffset']
            })
        
        models[model_name] = {
            'name': model_name,
            'description': f'{model_name} 设备模型',
            'protocol': 'ModbusTCP',
            'aiCount': template['aiCount'],
            'diCount': template['diCount'],
            'aiPoints': ai_points,
            'diPoints': di_points
        }
    
    # 保存结果
    result = {
        'templates': templates,
        'devices': devices_from_template,
        'sourceNodePoints': {k: len(v) for k, v in device_points.items()},
        'models': models
    }
    
    output_path = Path(file_path).parent / 'ism_data_models.json'
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    
    print(f"\n✅ 数据模型已保存到: {output_path}")
    print(f"\n模型摘要:")
    for name, model in models.items():
        print(f"  {name}: AI={len(model['aiPoints'])}, DI={len(model['diPoints'])}")
        print(f"    AI参数示例:")
        for ai in model['aiPoints'][:3]:
            print(f"      - {ai['name']}: 偏移={ai['offset']}, 系数={ai['factor']}, 解析方式={ai['parseMode']}")
    
    return result


if __name__ == '__main__':
    import sys
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
    else:
        file_path = '/Users/yunfanleo/cursorProjects/ISM源码/1A配电室 172.31.4.14 172.20.255.14.xlsx'
    
    analyze_excel(file_path)
