#!/usr/bin/env python3
"""
ISM 1A配电室数据生成脚本
解析 Excel 配置表 → 生成完整 _analysis.json + _device_points.json + ISM项目数据包 JSON
"""

import openpyxl
import json
import re
from collections import OrderedDict

EXCEL_PATH = "1A配电室 172.31.4.14 172.20.255.14.xlsx"

# ============================================================
# 1. 解析 Excel 模板 Sheet → 提取 3 种数据模型定义
# ============================================================

def parse_template_sheet(wb):
    """解析"模板"Sheet，提取A20/A40/UPS的完整寄存器定义"""
    ws = wb['模板']
    models = OrderedDict()
    current_model = None
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        col0 = str(row[0]).strip() if row[0] else ''
        
        if col0 and ('A20' in col0 or 'A40' in col0 or 'UPS' in col0 or '施耐德' in col0):
            current_model = col0.replace('\n', ' ')
            ai_count = 0; di_count = 0
            m_ai = re.search(r'AI\s*(\d+)', current_model)
            m_di = re.search(r'DI\s*(\d+)', current_model)
            if m_ai: ai_count = int(m_ai.group(1))
            if m_di: di_count = int(m_di.group(1))
            
            # Normalize model key
            if 'A20' in current_model: model_key = 'A20电力仪表'
            elif 'A40' in current_model: model_key = 'A40电力仪表'
            else: model_key = '施耐德UPS'
            
            models[model_key] = {
                'model_name': current_model,
                'ai_count': ai_count, 'di_count': di_count,
                'ai_points': [], 'di_points': [], 'devices': [],
            }
            continue
        
        if not current_model:
            continue
        
        # Normalize model key
        if 'A20' in current_model: model_key = 'A20电力仪表'
        elif 'A40' in current_model: model_key = 'A40电力仪表'
        else: model_key = '施耐德UPS'
        
        if model_key not in models:
            continue
        
        ai_name = str(row[1]).strip() if row[1] else ''
        ai_offset = row[2]; ai_coeff = row[3]; ai_parse = row[4]
        di_name = str(row[8]).strip() if row[8] else ''
        di_offset = row[9]; di_bit = row[10]
        dev_name = str(row[11]).strip() if row[11] else ''
        ai_start = row[14] if row[14] else None
        di_start = row[15] if row[15] else None
        
        if ai_name and ai_name not in ['None', '']:
            coeff_val = ai_coeff; parse_val = ai_parse
            # Detect swapped coeff/parse (A40 case)
            if isinstance(ai_coeff, (int, float)) and isinstance(ai_parse, (int, float)):
                if ai_coeff > 50 and ai_parse < 1:
                    parse_val = ai_coeff; coeff_val = ai_parse
            
            models[model_key]['ai_points'].append({
                'offset': int(ai_offset) if ai_offset is not None else None,
                'name': ai_name,
                'coeff': coeff_val if coeff_val else 1,
                'parse': int(parse_val) if parse_val else 177,
            })
        
        if di_name and di_name not in ['None', '']:
            models[model_key]['di_points'].append({
                'offset': int(di_offset) if di_offset is not None else None,
                'name': di_name,
                'bit_offset': int(di_bit) if di_bit is not None else None,
            })
        
        if dev_name and dev_name not in ['None', '']:
            models[model_key]['devices'].append({
                'device_name': dev_name,
                'ai_start': int(ai_start) if ai_start else None,
                'di_start': int(di_start) if di_start else None,
            })
    
    for key in list(models.keys()):
        models[key]['devices'] = [d for d in models[key]['devices'] if d['ai_start'] is not None]
    
    return models


# ============================================================
# 2. 解析主数据 Sheet → 提取 2858 行完整寄存器映射
# ============================================================

def parse_main_sheet(wb):
    """解析主数据Sheet"""
    ws = wb['1A配电室 172.31.4.14 172.20.255.14']
    records = []
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        point_no = row[0]; node_id = row[1]; node_point = row[2]
        node_name = str(row[3]).strip() if row[3] else ''
        point_name = str(row[4]).strip() if row[4] else ''
        reg_addr = row[5]
        
        # Skip non-data rows: headers, empty rows, etc.
        if point_no is None or not str(point_no).strip():
            continue
        try:
            pn = int(float(str(point_no)))
        except (ValueError, TypeError):
            continue
            
        try:
            ni = int(float(str(node_id))) if node_id is not None and str(node_id).strip() else None
        except (ValueError, TypeError):
            ni = str(node_id) if node_id else None
            
        try:
            np_val = int(float(str(node_point))) if node_point is not None and str(node_point).strip() else None
        except (ValueError, TypeError):
            np_val = None
            
        try:
            ra = int(float(str(reg_addr))) if reg_addr is not None and str(reg_addr).strip() else None
        except (ValueError, TypeError):
            ra = None
            
        records.append({
            'point_no': pn,
            'node_id': ni,
            'node_point': np_val,
            'node_name': node_name, 'point_name': point_name,
            'reg_addr': ra,
        })
    
    return records


# ============================================================
# 3. 解析设备清单
# ============================================================

def parse_device_sheets(wb):
    """解析 Sheet1 和 Sheet3"""
    devices_full = []
    
    # Sheet1
    ws1 = wb['Sheet1']
    for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row, values_only=True):
        full_name = str(row[0]).strip() if row[0] else ''
        short_name = str(row[1]).strip() if row[1] else ''
        ai_start = row[2] if row[2] else None
        if full_name and short_name:
            devices_full.append({
                'full_name': full_name, 'short_name': short_name,
                'ai_start': int(ai_start) if ai_start else None, 'di_start': None,
            })
    
    # Sheet3
    ws3 = wb['Sheet3']
    sheet3_data = {}
    for row in ws3.iter_rows(min_row=1, max_row=ws3.max_row, values_only=True):
        dev_name = str(row[0]).strip() if row[0] else ''
        ai_start = row[1] if row[1] else None; di_start = row[2] if row[2] else None
        if dev_name:
            sheet3_data[dev_name] = {
                'ai_start': int(ai_start) if ai_start else None,
                'di_start': int(di_start) if di_start else None,
            }
    
    for dev in devices_full:
        for s3_name, s3_info in sheet3_data.items():
            if dev['full_name'] in s3_name or s3_name.startswith(dev['full_name']):
                dev['di_start'] = s3_info['di_start']
                break
    
    return devices_full


# ============================================================
# 4. 工具函数
# ============================================================

def determine_device_type(full_name):
    if 'A40' in full_name.lower() or '_A40' in full_name: return 'A40电力仪表'
    if 'UPS' in full_name.lower() or '施耐德' in full_name: return '施耐德UPS'
    return 'A20电力仪表'


def get_unit_for_point(point_name):
    if '电压' in point_name: return 'V'
    if '电流' in point_name and '谐波' not in point_name: return 'A'
    if '功率' in point_name:
        if '因数' in point_name: return ''
        if '有功' in point_name or '视在' in point_name: return 'kW'
        if '无功' in point_name: return 'kvar'
        return 'kW'
    if '频率' in point_name: return 'Hz'
    if '电度' in point_name: return 'kWh'
    if '谐波畸变率' in point_name: return '%'
    if '温度' in point_name: return '°C'
    if '时间' in point_name: return 'min'
    return ''


def get_parse_label(code):
    return {177:'uint16', 179:'int32', 73:'int16', 71:'uint16'}.get(code, f'code_{code}')


def infer_location(short_name):
    full = short_name
    loc = {'building': '1A配电室', 'cabinet': '', 'unit': ''}
    for pat, cb in [(r'(1A\d)_U(\d+)_S(\d+)_(\d+)', 'U{n1}_S{n2}'),
                     (r'(1A\d)_U(\d+)_D(\d+)_(\d+)', 'U{n1}_D{n2}'),
                     (r'UPS_(1A\d)_U(\d+)', 'UPS_U{n1}')]:
        m = re.search(pat, full)
        if m:
            loc['building'] = f"{m.group(1)}区"
            loc['cabinet'] = cb.format(n1=m.group(2), n2=m.group(3) if len(m.groups()) > 3 else m.group(2))
            loc['unit'] = m.group(m.lastindex or 1)
            break
    loc['description'] = f"{loc['building']}, {loc['cabinet']}"
    return loc


# ============================================================
# MAIN
# ============================================================

def main():
    print("=" * 60)
    print("ISM 1A配电室数据完整生成")
    print("=" * 60)
    
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    
    # Step 1: Parse models
    print("\n[1/4] 解析数据模型模板...")
    models = parse_template_sheet(wb)
    for key, m in models.items():
        print(f"  {key}: AI={len(m['ai_points'])}点, DI={len(m['di_points'])}点")
    
    # Step 2: Parse devices
    print("\n[2/4] 解析设备清单...")
    devices_full = parse_device_sheets(wb)
    print(f"  共 {len(devices_full)} 台设备")
    
    # Step 3: Parse main data
    print("\n[3/4] 解析主数据表...")
    records = parse_main_sheet(wb)
    print(f"  共 {len(records)} 条记录")
    
    # Group devices and map data points
    devices_by_group = OrderedDict()
    records_by_node = OrderedDict()
    
    for rec in records:
        nn = rec['node_name']
        if nn not in records_by_node:
            records_by_node[nn] = []
        records_by_node[nn].append(rec)
    
    for dev in devices_full:
        model_type = determine_device_type(dev['full_name'])
        gm = re.match(r'(P\d+)', dev['full_name'])
        group = gm.group(1) if gm else '管理机'
        if group not in devices_by_group:
            devices_by_group[group] = []
        devices_by_group[group].append({**dev, 'model_type': model_type, 'data_points': []})
    
    # Map data points
    total_points = 0
    for group, devices in devices_by_group.items():
        # Find matching records
        candidates = []
        for nn, recs in records_by_node.items():
            if group in nn and (nn.startswith(f"{group}_") or f"_{group}" in nn):
                candidates = recs
                break
        
        if not candidates:
            for nn, recs in records_by_node.items():
                if group in nn:
                    candidates = recs
                    break
        
        if candidates:
            # Sequential distribution
            rec_idx = 0
            for dev in devices:
                model = models.get(dev['model_type'])
                if not model: continue
                pts_per = len(model['ai_points']) + len(model['di_points'])
                dev_recs = candidates[rec_idx:rec_idx + pts_per]
                if len(dev_recs) == 0: continue
                rec_idx += pts_per
                
                dev_idx = devices.index(dev)
                for pt_idx, dr in enumerate(dev_recs):
                    # Match to model point by POSITION within the device's data points
                    # First AI points, then DI points
                    model_pt = None
                    pt_type = 'AI'
                    ai_list = model['ai_points']
                    di_list = model['di_points']
                    
                    if pt_idx < len(ai_list):
                        model_pt = ai_list[pt_idx]
                        pt_type = 'AI'
                    elif pt_idx < len(ai_list) + len(di_list):
                        model_pt = di_list[pt_idx - len(ai_list)]
                        pt_type = 'DI'
                    
                    model_name = model_pt['name'] if model_pt else dr['point_name']
                    display_name = f"{model_name}"
                    
                    devices_by_group[group][dev_idx]['data_points'].append({
                        'name': dr['point_name'],
                        'display_name': display_name,
                        'model_name': model_name,
                        'point_type': pt_type,
                        'register_addr': dr['reg_addr'],
                        'model_offset': model_pt['offset'] if model_pt else None,
                        'coeff': model_pt.get('coeff', 1) if model_pt else 1,
                        'parse': model_pt.get('parse', 177) if model_pt else 177,
                        'parse_label': get_parse_label(model_pt.get('parse', 177)) if model_pt else 'uint16',
                        'unit': get_unit_for_point(model_name),
                        'conversion_expression': f"x*{model_pt.get('coeff', 1)}" if model_pt else "x",
                        'is_alarm': '通讯状态' in model_name,
                        'alarm_level': 3 if '通讯状态' in model_name else 0,
                        'alarm_message': '设备通讯离线' if '通讯状态' in model_name else '',
                    })
                total_points += len(dev_recs)
    
    print(f"  总计: {sum(len(d) for d in devices_by_group.values())}台设备, {total_points}个数据点")
    
    # ==============================================
    # OUTPUT 1: _analysis.json
    # ==============================================
    print("\n[4/4] 生成输出文件...")
    
    analysis = OrderedDict()
    analysis["project_name"] = "1A配电室"
    analysis["protocol"] = "Modbus TCP"
    analysis["gateway"] = {"primary_ip": "172.31.4.14", "secondary_ip": "172.20.255.14", "port": 502}
    analysis["total_devices"] = len(devices_full)
    analysis["total_data_points"] = total_points
    analysis["device_categories"] = OrderedDict()
    
    cat_configs = {
        "A20电力仪表": (30, 3, 30),
        "A40电力仪表": (41, 3, 44),
        "施耐德UPS":    (44, 1, 45),
        "管理机(装置本身)": (0, 76, 0),
    }
    
    for cat_key, (ai_n, di_n, reg_off) in cat_configs.items():
        cat = OrderedDict()
        cat["description"] = f"{ai_n}个AI寄存器 + {di_n}个DI寄存器"
        cat["ai_count"] = ai_n
        cat["di_count"] = di_n
        cat["register_offset_per_device"] = reg_off
        cat["devices"] = []
        cat["data_point_types"] = OrderedDict()
        
        if cat_key in models:
            m = models[cat_key]
            cat["data_point_types"]["AI"] = [
                OrderedDict([
                    ("offset", pt["offset"]), ("name", pt["name"]),
                    ("coeff", pt["coeff"]), ("parse", pt["parse"]),
                    ("parse_label", get_parse_label(pt["parse"])),
                    ("unit", get_unit_for_point(pt["name"])),
                ]) for pt in m["ai_points"]
            ]
            cat["data_point_types"]["DI"] = [
                OrderedDict([
                    ("offset", pt["offset"]), ("name", pt["name"]),
                    ("bit_offset", pt["bit_offset"]),
                ]) for pt in m["di_points"]
            ]
        
        # Fill devices
        for group, devs in devices_by_group.items():
            for d in devs:
                if d["model_type"] == cat_key:
                    cat["devices"].append(OrderedDict([
                        ("group", group), ("full_name", d["full_name"]),
                        ("short_name", d["short_name"]),
                        ("ai_start", d["ai_start"]), ("di_start", d["di_start"]),
                        ("data_point_count", len(d["data_points"])),
                    ]))
        
        # Management device
        if cat_key == "管理机(装置本身)":
            gw_rn = "1A配电室 172.31.4.14 172.20.255.14"
            if gw_rn in records_by_node:
                cat["data_point_types"]["DIO"] = [
                    OrderedDict([("name", r["point_name"]), ("register_addr", r["reg_addr"])])
                    for r in records_by_node[gw_rn]
                ]
        
        analysis["device_categories"][cat_key] = cat
    
    with open("1A配电室_analysis.json", "w", encoding="utf-8") as f:
        json.dump(analysis, f, ensure_ascii=False, indent=2)
    print("  ✓ 1A配电室_analysis.json")
    
    # ==============================================
    # OUTPUT 2: _device_points.json
    # ==============================================
    
    dpts = OrderedDict()
    dpts["project"] = "1A配电室"
    dpts["gateway_ip"] = "172.31.4.14"
    dpts["gateway_port"] = 502
    dpts["total_devices"] = len(devices_full)
    dpts["total_data_points"] = total_points
    dpts["device_groups"] = OrderedDict()
    
    for group, devs in devices_by_group.items():
        all_pts = []
        for d in devs:
            all_pts += d["data_points"]
        dpts["device_groups"][group] = OrderedDict([
            ("model_type", devs[0]["model_type"] if devs else "Unknown"),
            ("device_count", len(devs)),
            ("point_count", len(all_pts)),
            ("devices", [
                OrderedDict([
                    ("full_name", d["full_name"]), ("short_name", d["short_name"]),
                    ("ai_start", d["ai_start"]), ("di_start", d["di_start"]),
                    ("data_points", d["data_points"]),
                ]) for d in devs
            ]),
        ])
    
    # Gateway
    gw_rn = "1A配电室 172.31.4.14 172.20.255.14"
    if gw_rn in records_by_node:
        dpts["device_groups"]["管理机(网关)"] = OrderedDict([
            ("model_type", "管理机(装置本身)"),
            ("device_count", 1), ("point_count", len(records_by_node[gw_rn])),
            ("devices", [OrderedDict([
                ("full_name", gw_rn), ("short_name", "管理机"),
                ("data_points", [
                    OrderedDict([("name", r["point_name"]), ("register_addr", r["reg_addr"])])
                    for r in records_by_node[gw_rn]
                ]),
            ])]),
        ])
    
    with open("1A配电室_device_points.json", "w", encoding="utf-8") as f:
        json.dump(dpts, f, ensure_ascii=False, indent=2)
    print("  ✓ 1A配电室_device_points.json")
    
    # ==============================================
    # OUTPUT 3: ISM项目包
    # ==============================================
    
    data_models = []
    for cat_key in ["A20电力仪表", "A40电力仪表", "施耐德UPS"]:
        m = models.get(cat_key)
        if not m: continue
        
        reg_groups = []
        # AI group
        ai_addrs = []
        for pt in m["ai_points"]:
            addr = OrderedDict([
                ("name", pt["name"]), ("offset", pt["offset"]),
                ("register_type", "holding"),
                ("data_type", get_parse_label(pt["parse"])),
                ("unit", get_unit_for_point(pt["name"])),
                ("coeff", pt["coeff"]),
                ("conversion_expression", f"x*{pt['coeff']}"),
                ("is_alarm", False), ("is_record", True),
            ])
            ai_addrs.append(addr)
        reg_groups.append(OrderedDict([("group_name", "AI模拟量"), ("register_type", "holding"), ("addresses", ai_addrs)]))
        
        # DI group
        di_addrs = []
        for pt in m["di_points"]:
            is_comm = "通讯状态" in pt["name"]
            addr = OrderedDict([
                ("name", pt["name"]), ("offset", pt["offset"]),
                ("register_type", "coil"), ("data_type", "bit"),
                ("unit", ""), ("coeff", 1),
                ("is_alarm", is_comm),
                ("alarm_level", 3 if is_comm else 0),
                ("alarm_message", f"设备通讯离线" if is_comm else ""),
            ])
            di_addrs.append(addr)
        reg_groups.append(OrderedDict([("group_name", "DI数字量"), ("register_type", "coil"), ("addresses", di_addrs)]))
        
        data_models.append(OrderedDict([
            ("name", cat_key), ("protocol", "modbus"),
            ("connection", OrderedDict([("ip", "172.31.4.14"), ("port", 502), ("slave_id", 1)])),
            ("register_groups", reg_groups),
        ]))
    
    # Devices
    ism_devices = []
    for group, devs in devices_by_group.items():
        for d in devs:
            loc = infer_location(d["short_name"])
            ism_devices.append(OrderedDict([
                ("name", d["short_name"]), ("display_name", d["full_name"]),
                ("model_name", d["model_type"]), ("group", group),
                ("register_offset", d["ai_start"]), ("di_start", d["di_start"]),
                ("location", loc),
                ("data_point_count", len(d["data_points"])),
            ]))
    
    # Alarms
    alarms = []
    for d in ism_devices:
        if d["model_name"] == "管理机(装置本身)": continue
        alarms.append(OrderedDict([
            ("name", f"{d['name']}_通讯离线告警"),
            ("device_name", d["name"]), ("data_point", "主通讯状态"),
            ("condition", "x == 0"), ("level", 3), ("keep_time", 10),
            ("alarm_message", f"{d['name']} 设备通讯离线"),
        ]))
    
    # Dashboard
    total_devs = sum(len(ds) for ds in devices_by_group.values())
    cabinets_1a1 = OrderedDict()
    cabinets_1a3 = OrderedDict()
    
    for group, devs in devices_by_group.items():
        for d in devs:
            loc = infer_location(d["short_name"])
            cab = loc["cabinet"]
            if "1A3" in loc["building"]:
                if cab not in cabinets_1a3: cabinets_1a3[cab] = []
                cabinets_1a3[cab].append(d)
            else:
                if cab not in cabinets_1a1: cabinets_1a1[cab] = []
                cabinets_1a1[cab].append(d)
    
    dashboard = OrderedDict([
        ("name", "1A配电室监控总览"),
        ("size", "1920x1080"),
        ("style", OrderedDict([("theme", "科技蓝"), ("primary_color", "#00d4ff"), ("bg_color", "#0a1628")])),
        ("overview_page", OrderedDict([
            ("name", "总览"), ("is_home", True),
            ("sections", [
                OrderedDict([
                    ("name", "顶部"), ("zone", "top"), ("height", 80),
                    ("component", OrderedDict([("type", "DvBorderBox1"), ("text", f"1A配电室 - 监控总览 ({total_devs}台设备)")])),
                ]),
                OrderedDict([
                    ("name", "左侧设备树"), ("zone", "left"), ("width", 280),
                    ("component", OrderedDict([("type", "DeviceTree"), ("title", "设备拓扑")])),
                ]),
                OrderedDict([
                    ("name", "1A1区配电柜"), ("zone", "center_top"), ("height", 480),
                    ("description", f"共 {sum(len(v) for v in cabinets_1a1.values())} 台设备, {len(cabinets_1a1)} 个配电柜"),
                    ("cabinets", [
                        OrderedDict([
                            ("name", cn), ("device_count", len(cd)),
                            ("devices", [OrderedDict([
                                ("name", dd["short_name"]), ("full_name", dd["full_name"]),
                                ("model_type", dd["model_type"]),
                                ("key_data", ["AB线电压", "A相电流", "总有功功率", "主通讯状态"]),
                                ("interactions", {
                                    "hover": {"show": f"detail_popup_{dd['short_name']}"},
                                    "click": {"navigate": f"cabinet_{cn}"},
                                }),
                            ]) for dd in cd]),
                        ]) for cn, cd in cabinets_1a1.items()
                    ]),
                ]),
                OrderedDict([
                    ("name", "1A3区配电柜"), ("zone", "center_bottom"), ("height", 480),
                    ("description", f"共 {sum(len(v) for v in cabinets_1a3.values())} 台设备, {len(cabinets_1a3)} 个配电柜"),
                    ("cabinets", [
                        OrderedDict([
                            ("name", cn), ("device_count", len(cd)),
                            ("devices", [OrderedDict([
                                ("name", dd["short_name"]), ("full_name", dd["full_name"]),
                                ("model_type", dd["model_type"]),
                                ("key_data", ["AB线电压", "A相电流", "总有功功率", "主通讯状态"]),
                            ]) for dd in cd]),
                        ]) for cn, cd in cabinets_1a3.items()
                    ]),
                ]),
                OrderedDict([
                    ("name", "右侧告警"), ("zone", "right"), ("width", 660),
                    ("components", [
                        OrderedDict([("type", "alarmList"), ("title", "实时告警")]),
                        OrderedDict([("type", "RealDataTable"), ("title", "关键数据"), ("columns", ["设备", "AB线电压", "A相电流", "总有功功率"])]),
                    ]),
                ]),
            ]),
        ])),
        ("interaction_templates", OrderedDict([
            ("hover_show_detail", OrderedDict([
                ("trigger", "mouseenter"),
                ("action", "visible"),
                ("showItems", "[detail_popup_{{device_name}}]"),
                ("hide_on", OrderedDict([("trigger", "mouseleave"), ("action", "visible"), ("hideItems", "[detail_popup_{{device_name}}]")])),
            ])),
            ("click_navigate", OrderedDict([
                ("trigger", "click"),
                ("action", "link"),
                ("linkType", "Inside"),
                ("target", "{{dashboard_uid}}:{{page_uid}}"),
            ])),
            ("click_device_view", OrderedDict([
                ("trigger", "click"),
                ("action", "DeviceView"),
                ("key", "{{device_uuid}}"),
            ])),
        ])),
    ])
    
    pkg = OrderedDict([
        ("_meta", OrderedDict([
            ("format_version", "1.0"), ("generator", "ISM AI Agent Data Generator v2"),
            ("source", EXCEL_PATH), ("generated_at", "2026-06-12"),
        ])),
        ("project", OrderedDict([
            ("name", "1A配电室"), ("description", "1A配电室 Modbus TCP 监控"),
            ("gateway", OrderedDict([
                ("primary_ip", "172.31.4.14"), ("secondary_ip", "172.20.255.14"), ("port", 502),
            ])),
        ])),
        ("data_models", data_models),
        ("devices", ism_devices),
        ("alarm_triggers", alarms),
        ("dashboard", dashboard),
        ("import_guide", [
            "1. 创建项目 → POST /api/project/add",
            "2. 导入3个数据模型 → POST /api/v1/open/models/batch",
            "3. 导入{len(ism_devices)}台设备 → POST /api/v1/open/devices/batch",
            "4. 导入{len(alarms)}个告警触发器 → POST /api/v1/open/alarms/batch",
            "5. 生成组态大屏 → MCP: dashboard_generate",
            "6. 一键导入全部 → MCP: import_project_from_excel",
        ]),
    ])
    
    with open("1A配电室_ISM项目包.json", "w", encoding="utf-8") as f:
        json.dump(pkg, f, ensure_ascii=False, indent=2)
    print("  ✓ 1A配电室_ISM项目包.json")
    
    print("\n" + "=" * 60)
    print("全部完成！生成了 3 个文件：")
    print("  1. 1A配电室_analysis.json - 数据模型分析（含完整AI/DI定义）")
    print("  2. 1A配电室_device_points.json - 设备点位明细（75设备+数据点）")
    print("  3. 1A配电室_ISM项目包.json - 可导入ISM的完整项目包")
    print("=" * 60)


if __name__ == "__main__":
    main()
